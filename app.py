from flask import Flask, render_template, request, jsonify, send_file
from flask_httpauth import HTTPBasicAuth
from flask_caching import Cache
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, time
import os
import pandas as pd
import numpy as np
from dotenv import load_dotenv
import hashlib
import json
import itertools
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re

load_dotenv()

# Intentar importar pyathena, pero permitir que la app funcione sin él
try:
    from pyathena import connect
    from pyathena.pandas.cursor import PandasCursor
    ATHENA_AVAILABLE = True
except ImportError:
    ATHENA_AVAILABLE = False
    print("⚠️ pyathena no disponible. La app usará datos de ejemplo de SQLite.")

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', '')

# ============================================
# CONFIGURACIÓN DE CACHÉ
# ============================================
# Configurar cache - usar SimpleCache para desarrollo, Redis para producción
cache_config = {
    'CACHE_TYPE': os.getenv('CACHE_TYPE', 'SimpleCache'),  # 'SimpleCache' o 'RedisCache'
    'CACHE_DEFAULT_TIMEOUT': int(os.getenv('CACHE_TIMEOUT', '900')),  # 15 minutos por defecto
}

# Si se usa Redis, agregar configuración
if cache_config['CACHE_TYPE'] == 'RedisCache':
    cache_config.update({
        'CACHE_REDIS_URL': os.getenv('REDIS_URL', 'redis://localhost:6379/0')
    })

app.config.update(cache_config)
cache = Cache(app)

print(f"📦 Caché configurado: {cache_config['CACHE_TYPE']} (timeout: {cache_config['CACHE_DEFAULT_TIMEOUT']}s)")

# ============================================
# UTILIDAD PARA LIMPIAR NaN EN JSON
# ============================================
def clean_nan_for_json(obj):
    """
    Recursivamente limpia valores NaN/inf para que sean JSON-serializables.
    Convierte NaN, inf, -inf a None.
    """
    if isinstance(obj, dict):
        return {k: clean_nan_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_nan_for_json(item) for item in obj]
    elif isinstance(obj, float):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return obj
    elif pd.isna(obj):
        return None
    return obj

# ============================================
# MIDDLEWARE PARA MEDICIÓN DE TIEMPOS
# ============================================
@app.before_request
def before_request():
    """Registrar tiempo de inicio de request"""
    from flask import g
    import time
    g.start_time = time.time()

@app.after_request
def after_request(response):
    """Registrar tiempo de respuesta"""
    from flask import g, request
    import time
    
    if hasattr(g, 'start_time'):
        elapsed = time.time() - g.start_time
        if request.path.startswith('/api/'):
            print(f"⏱️  {request.method} {request.path} - {elapsed:.2f}s")
    
    return response

# ============================================
# CONFIGURACIÓN DE AUTENTICACIÓN
# ============================================
auth = HTTPBasicAuth()

# Usuarios permitidos (desde variables de entorno)
USERS = {
    os.environ.get('APP_USER'): generate_password_hash(
        os.environ.get('APP_PASSWORD')
    )
}

@auth.verify_password
def verify_password(username, password):
    """Verificar credenciales de usuario"""
    if username in USERS and check_password_hash(USERS.get(username), password):
        return username
    return None
# ============================================
# FIN CONFIGURACIÓN DE AUTENTICACIÓN
# ============================================


# Configuración de AWS Athena (usar variables de entorno en producción)
AWS_CONFIG = {
    'aws_access_key_id': os.getenv('AWS_ACCESS_KEY', ''),
    'aws_secret_access_key': os.getenv('AWS_SECRET_KEY', ''),
    's3_staging_dir': os.getenv('S3_BUCKET', ''),
    'region_name': os.getenv('AWS_REGION', '')
}

# Nombre base de datos/tabla en Athena (desde variables de entorno)
DATABASE_ATHENA = os.getenv('DATABASE_ATHENA', '')
TABLA_PEDIDOS = os.getenv('TABLA_PEDIDOS', '')

# Configuración adicional para el análisis completo
DATABASE_DISPOMATE = os.getenv('DATABASE_DISPOMATE', '')
TABLA_TURNOS = os.getenv('TABLA_TURNOS', '')
TABLA_RCO = os.getenv('TABLA_RCO', '')
TABLA_FLOTA = os.getenv('TABLA_FLOTA', '')


# Forzar por defecto el uso de Athena si la librería está disponible.
# Si pyathena no está instalado, USE_ATHENA será False y la app informará al usuario.
USE_ATHENA = os.getenv('USE_ATHENA', 'true').lower() == 'true' and ATHENA_AVAILABLE

# Conexión global a Athena
athena_conn = None

# ============================================
# FUNCIONES HELPER PARA CACHÉ
# ============================================
def generar_cache_key(prefix, *args, **kwargs):
    """Generar clave única para cache basada en parámetros"""
    key_data = f"{prefix}_{args}_{sorted(kwargs.items())}"
    key_hash = hashlib.md5(key_data.encode()).hexdigest()[:12]
    return f"{prefix}_{key_hash}"

def invalidar_cache_minera(minera_nombre):
    """Invalidar todo el cache relacionado con una minera específica"""
    cache.clear()
    print(f"🗑️ Cache invalidado para minera: {minera_nombre}")

# ============================================
# CONEXIÓN A ATHENA
# ============================================
def get_athena_connection():
    """Obtener o crear conexión a Athena"""
    global athena_conn
    if not USE_ATHENA or not ATHENA_AVAILABLE:
        return None
    
    if athena_conn is None:
        try:
            athena_conn = connect(
                cursor_class=PandasCursor,
                **AWS_CONFIG
            )
            print("✅ Conexión a Athena establecida")
        except Exception as e:
            print(f"❌ Error conectando a Athena: {e}")
            return None
    return athena_conn

def sql_athena(query):
    """Ejecutar query en Athena y retornar DataFrame"""
    try:
        conn = get_athena_connection()
        if conn:
            return conn.cursor().execute(query).as_pandas()
        raise RuntimeError("Athena no disponible: no se puede ejecutar queries")
    except Exception as e:
        print(f"Error ejecutando query en Athena: {e}")
        return pd.DataFrame()


# ============================================
# PARSER DE TURNOS DESDE MANTENEDOR DE FLOTA
# ============================================

def parsear_turnos_detalle(turnos_detalle_str):
    """
    Parsear la columna 'Turnos detalle' del mantenedor de flota.
    
    Formato esperado:
    - Un turno: "07:00:00 - Lunes,Martes,Miercoles,Jueves,Viernes,Sábado"
    - Múltiples turnos (separados por salto de línea o \\n):
      "07:00:00 - Domingo,Lunes,Martes,Miercoles,Jueves,Viernes,Sábado
       19:00:00 - Domingo,Lunes,Martes,Miercoles,Jueves,Viernes,Sábado"
    
    Returns:
        List[dict]: Lista de turnos con estructura:
            [
                {
                    'hora_inicio': '07:00:00',
                    'dias': ['Lunes', 'Martes', ...],
                    'turno_nombre': 'Turno AM'  # o 'Turno PM', 'Turno Noche'
                },
                ...
            ]
    """
    if pd.isna(turnos_detalle_str) or not turnos_detalle_str or turnos_detalle_str.strip() == '':
        return []
    
    # Limpiar el string y separar por saltos de línea (pueden ser \n o literales)
    turnos_str = str(turnos_detalle_str).strip()
    
    # Manejar strings con comillas dobles que envuelven múltiples líneas
    if turnos_str.startswith('"') and turnos_str.endswith('"'):
        turnos_str = turnos_str[1:-1]
    
    # Separar por saltos de línea (tanto \n como saltos literales)
    lineas_turnos = [linea.strip() for linea in re.split(r'\n|\\n', turnos_str) if linea.strip()]
    
    turnos_parseados = []
    
    for idx, linea in enumerate(lineas_turnos):
        # Formato: "HH:MM:SS - Dia1,Dia2,Dia3..."
        partes = linea.split(' - ', 1)
        
        if len(partes) != 2:
            continue
        
        hora_inicio = partes[0].strip()
        dias_str = partes[1].strip()
        
        # Separar días
        dias = [dia.strip() for dia in dias_str.split(',')]
        
        # Usar la hora directamente como nombre del turno (formato HH:MM)
        try:
            hora_obj = datetime.strptime(hora_inicio, '%H:%M:%S').time()
            turno_nombre = hora_obj.strftime('%H:%M')  # Formato 07:00, 19:00, etc.
        except:
            # Si no se puede parsear la hora, usar la hora tal cual o índice
            turno_nombre = hora_inicio if hora_inicio else f'Turno {idx + 1}'
        
        turnos_parseados.append({
            'hora_inicio': hora_inicio,
            'dias': dias,
            'turno_nombre': turno_nombre
        })
    
    return turnos_parseados

@cache.memoize(timeout=3600)  # Cache por 1 hora
def cargar_mantenedor_flota():
    """
    Cargar el archivo mantenedor_flota.txt y parsearlo.
    
    Returns:
        pd.DataFrame con columnas:
            - Equipo (código del camión)
            - Transportista
            - Turnos (lista de diccionarios con información de turnos)
    """
    print("🔄 [CACHE MISS] Cargando mantenedor de flota...")
    
    # Ruta al archivo (ajustar según tu configuración)
    rutas_txt_posibles = [
        '/etc/secrets/mantenedor_flota.txt',  # Render Secret Files (ubicación alternativa)
        os.path.join(os.path.dirname(__file__), 'mantenedor_flota.txt'),  # Raíz de la app (Render o local)
        os.path.join(os.path.dirname(__file__), 'Flota_20260108 - copia.txt')  # Nombre alternativo local
    ]
    
    ruta_mantenedor = None
    for ruta in rutas_txt_posibles:
        if os.path.exists(ruta):
            ruta_mantenedor = ruta
            break
    
    if not os.path.exists(ruta_mantenedor):
        print(f"⚠️ Archivo mantenedor no encontrado: {ruta_mantenedor}")
        return pd.DataFrame()
    
    try:
        # Leer archivo delimitado por tabuladores
        df_flota = pd.read_csv(ruta_mantenedor, sep='\t', encoding='latin-1')
        
        # Parsear turnos para cada vehículo
        df_flota['Turnos_parseados'] = df_flota['Turnos detalle'].apply(parsear_turnos_detalle)
        
        # Quedarnos solo con las columnas relevantes
        df_flota_limpio = df_flota[['Equipo', 'Transportista', 'Turnos_parseados']].copy()
        df_flota_limpio = df_flota_limpio.rename(columns={'Equipo': 'Camion'})
        
        print(f"✅ Mantenedor cargado: {len(df_flota_limpio)} vehículos")
        
        return df_flota_limpio
        
    except Exception as e:
        print(f"❌ Error cargando mantenedor de flota: {e}")
        return pd.DataFrame()

def expandir_turnos_a_registros(df_base, df_mantenedor):
    """
    Expandir un DataFrame de [Camión, Fecha] a [Camión, Fecha, Turno]
    usando la información del mantenedor de flota.
    
    Crea un registro por cada turno activo en el día, permitiendo ver
    todos los turnos incluso si no tienen actividad.
    
    Args:
        df_base: DataFrame con estructura [Camion, fecha_completa, ...]
        df_mantenedor: DataFrame del mantenedor con turnos parseados
    
    Returns:
        DataFrame expandido con una fila por cada turno activo
    """
    print("🔄 Expandiendo registros a nivel de turno...")
    
    # Mapeo de días en español a inglés
    dias_semana_map = {
        'Lunes': 'Monday',
        'Martes': 'Tuesday',
        'Miercoles': 'Wednesday',
        'Miércoles': 'Wednesday',
        'Jueves': 'Thursday',
        'Viernes': 'Friday',
        'Sabado': 'Saturday',
        'Sábado': 'Saturday',
        'Domingo': 'Sunday'
    }
    
    registros_expandidos = []
    
    # Iterar por cada camión en df_base
    for camion in df_base['Camion'].unique():
        # Obtener configuración de turnos del mantenedor
        config_camion = df_mantenedor[df_mantenedor['Camion'] == camion]
        
        if config_camion.empty:
            # Si no hay configuración, mantener el registro sin turno
            df_camion = df_base[df_base['Camion'] == camion].copy()
            df_camion['Turno'] = 'Sin Turno'
            df_camion['Hora_inicio_turno'] = None
            registros_expandidos.append(df_camion)
            continue
        
        turnos_config = config_camion.iloc[0]['Turnos_parseados']
        
        if not turnos_config or len(turnos_config) == 0:
            # Sin turnos configurados
            df_camion = df_base[df_base['Camion'] == camion].copy()
            df_camion['Turno'] = 'Sin Turno'
            df_camion['Hora_inicio_turno'] = None
            registros_expandidos.append(df_camion)
            continue
        
        # Obtener datos del camión
        df_camion = df_base[df_base['Camion'] == camion].copy()
        
        # Para cada fecha del camión
        for _, row in df_camion.iterrows():
            fecha = row['fecha_completa']
            dia_semana = fecha.strftime('%A')  # Nombre del día en inglés
            
            # Encontrar qué turnos aplican a este día
            turnos_del_dia = []
            for turno_config in turnos_config:
                # Verificar si este turno opera en este día de la semana
                dias_turno_esp = turno_config['dias']
                
                # Convertir días del turno a inglés
                dias_turno_ing = []
                for dia_esp in dias_turno_esp:
                    dia_esp_norm = dia_esp.strip()
                    if dia_esp_norm in dias_semana_map:
                        dias_turno_ing.append(dias_semana_map[dia_esp_norm])
                
                if dia_semana in dias_turno_ing:
                    turnos_del_dia.append(turno_config)
            
            # Si hay turnos para este día, crear un registro por turno
            if turnos_del_dia:
                for turno_config in turnos_del_dia:
                    registro_turno = row.copy()
                    registro_turno['Turno'] = turno_config['turno_nombre']
                    registro_turno['Hora_inicio_turno'] = turno_config['hora_inicio']
                    registros_expandidos.append(pd.DataFrame([registro_turno]))
            else:
                # No hay turnos para este día, mantener sin turno
                registro_sin_turno = row.copy()
                registro_sin_turno['Turno'] = 'Sin Turno'
                registro_sin_turno['Hora_inicio_turno'] = None
                registros_expandidos.append(pd.DataFrame([registro_sin_turno]))
    
    if not registros_expandidos:
        return df_base.copy()
    
    # Concatenar todos los registros
    df_expandido = pd.concat(registros_expandidos, ignore_index=True)
    
    print(f"✅ Expansión completada: {len(df_base)} → {len(df_expandido)} registros")
    
    return df_expandido

# ============================================
# FUNCIÓN PRINCIPAL: OBTENER DATOS COMPLETOS CON TURNOS
# ============================================
@cache.memoize(timeout=900)  # Cache por 15 minutos
def obtener_datos_completos_athena(minera_nombre, fecha_inicio, fecha_fin):
    """
    Obtener datos completos integrando SCR, Turnos y RCO.
    NUEVA VERSIÓN SIMPLIFICADA: Queries directas sin merges complejos.
    
    Estructura de retorno:
        DataFrame con columnas:
        - Camion
        - Fecha
        - fecha_completa
        - Turno (ej: 'N/A' por ahora, sin granularidad de turno para evitar pérdida de datos)
        - Entregado_totalmente
        - En_ruta
        - Planificado
        - Turnos_enviados
        - Conexion_RCO
        - Transportista
        - ¿Es licitado?
    """
    print(f"🔄 [CACHE MISS] Ejecutando queries completas SIMPLIFICADAS para {minera_nombre} ({fecha_inicio} - {fecha_fin})")
    
    if not USE_ATHENA:
        return None

    # Obtener flota licitada para la minera
    df_flota = obtener_flota_licitada_athena(minera_nombre)
    
    if df_flota.empty:
        print("⚠️ No se encontró flota licitada para esta minera")
        return None
    
    vehiculos_transportista = df_flota['codigo_tanque'].tolist()
    print(f"📋 Total vehículos en flota: {len(vehiculos_transportista)}")
    
    # 1. QUERY TURNOS ENVIADOS (POR FECHA Y HORA SANTIAGO)
    # Restar 3 horas para convertir de UTC a Santiago (Chile Continental sin horario de verano)
    query_turnos = f"""
    SELECT 
        id_vehiculo, 
        DATE(fecha_inicio_utc - INTERVAL '3' HOUR) as fecha, 
        CAST(fecha_inicio_utc - INTERVAL '3' HOUR AS TIME) as hora_inicio_turno,
        COUNT(DISTINCT id_turno_uuid) as turnos_enviados
    FROM (
        SELECT id_turno_uuid, id_vehiculo, estado, fecha_inicio_utc,
               ROW_NUMBER() OVER (PARTITION BY id_turno_uuid ORDER BY fecha_hora DESC) as rn
        FROM {DATABASE_DISPOMATE}.{TABLA_TURNOS}
        WHERE DATE(fecha_inicio_utc - INTERVAL '3' HOUR) >= DATE('{fecha_inicio.strftime('%Y-%m-%d')}')
          AND DATE(fecha_inicio_utc - INTERVAL '3' HOUR) <= DATE('{fecha_fin.strftime('%Y-%m-%d')}')
          AND id_vehiculo IN ({','.join(map(str, vehiculos_transportista))})
    )
    WHERE rn = 1 AND estado = 'ENVIADO'
    GROUP BY id_vehiculo, DATE(fecha_inicio_utc - INTERVAL '3' HOUR), CAST(fecha_inicio_utc - INTERVAL '3' HOUR AS TIME)
    """
    
    try:
        df_turnos = sql_athena(query_turnos)
        print(f"📊 DEBUG TURNOS - Total registros: {len(df_turnos)}")
        if not df_turnos.empty:
            print(f"📊 DEBUG TURNOS - Columnas: {df_turnos.columns.tolist()}")
            print(f"📊 DEBUG TURNOS - Primeros 10 registros:")
            print(df_turnos[['id_vehiculo', 'fecha', 'hora_inicio_turno', 'turnos_enviados']].head(10))
    except Exception as e:
        print(f"Error obteniendo turnos: {e}")
        df_turnos = pd.DataFrame()
    
    # 2. QUERY CONEXIONES RCO (CON FECHA Y HORA PARA ASIGNACIÓN DE TURNO)
    query_rco = f"""
    SELECT 
        codigo_tanque,
        DATE(fecha_conexion_utc) as fecha,
        CAST(fecha_conexion_utc AS TIME) as hora,
        COUNT(*) as conexiones_rco
    FROM {DATABASE_DISPOMATE}.{TABLA_RCO}
    WHERE DATE(fecha_conexion_utc) >= DATE('{fecha_inicio.strftime('%Y-%m-%d')}')
      AND DATE(fecha_conexion_utc) <= DATE('{fecha_fin.strftime('%Y-%m-%d')}')
      AND codigo_tanque IN ({','.join(map(str, vehiculos_transportista))})
    GROUP BY codigo_tanque, DATE(fecha_conexion_utc), CAST(fecha_conexion_utc AS TIME)
    """
    
    try:
        df_rco = sql_athena(query_rco)
        if not df_rco.empty:
            # Convertir fecha a datetime
            df_rco['fecha'] = pd.to_datetime(df_rco['fecha']).dt.date
    except Exception as e:
        print(f"Error obteniendo RCO: {e}")
        df_rco = pd.DataFrame()
    
    # 3. QUERY ENTREGAS (CON FECHA Y HORA PARA ASIGNACIÓN DE TURNO)
    query_entregas = f"""
    SELECT 
        CAST(vehiclereal AS INTEGER) as vehiclereal,
        fechallegadaprog as fecha,
        horallegadaprog as hora,
        descrstatu as estado,
        COUNT(*) as cantidad
    FROM {DATABASE_ATHENA}.{TABLA_PEDIDOS}
    WHERE fechallegadaprog >= '{fecha_inicio.strftime('%Y-%m-%d')}'
      AND fechallegadaprog <= '{fecha_fin.strftime('%Y-%m-%d')}'
      AND CAST(vehiclereal AS INTEGER) IN ({','.join(map(str, vehiculos_transportista))})
    GROUP BY CAST(vehiclereal AS INTEGER), fechallegadaprog, horallegadaprog, descrstatu
    """
    
    try:
        df_entregas = sql_athena(query_entregas)
        if not df_entregas.empty:
            df_entregas['vehiclereal'] = df_entregas['vehiclereal'].astype(int)
            # Convertir fecha a datetime
            df_entregas['fecha'] = pd.to_datetime(df_entregas['fecha']).dt.date
            print(f"📊 DEBUG ENTREGAS - Total registros: {len(df_entregas)}")
            print(f"📊 DEBUG ENTREGAS - Fechas únicas: {sorted(df_entregas['fecha'].unique().tolist())}")
            print(f"📊 DEBUG ENTREGAS - Estados únicos: {df_entregas['estado'].unique().tolist()}")
            print(f"📊 DEBUG ENTREGAS - Camiones con entregas: {sorted(df_entregas['vehiclereal'].unique().tolist())}")
            
            # Contar por estado
            print(f"📊 DEBUG ENTREGAS - Cantidad por estado:")
            for estado in df_entregas['estado'].unique():
                cant = df_entregas[df_entregas['estado'] == estado]['cantidad'].sum()
                print(f"   {estado}: {cant}")
            
            print(f"📊 DEBUG ENTREGAS - Primeros 20 registros:")
            print(df_entregas[['vehiclereal', 'fecha', 'hora', 'estado', 'cantidad']].head(20))
        else:
            print("⚠️ DEBUG ENTREGAS - DataFrame vacío")
    except Exception as e:
        print(f"Error obteniendo entregas: {e}")
        df_entregas = pd.DataFrame()
    
    # 4. CREAR DICCIONARIO DE TURNOS POR VEHÍCULO Y FECHA
    # Ahora usamos la hora_inicio_turno de la tabla de turnos directamente
    turnos_por_vehiculo_fecha = {}
    if not df_turnos.empty and 'hora_inicio_turno' in df_turnos.columns:
        for _, row in df_turnos.iterrows():
            vehiculo = int(row['id_vehiculo'])
            fecha_str = row['fecha']
            
            # Convertir fecha a string en formato YYYY-MM-DD para consistencia
            if isinstance(fecha_str, pd.Timestamp):
                fecha_str = fecha_str.strftime('%Y-%m-%d')
            elif isinstance(fecha_str, datetime):
                fecha_str = fecha_str.strftime('%Y-%m-%d')
            else:
                fecha_str = str(fecha_str)
            
            hora_turno = str(row['hora_inicio_turno'])  # Formato HH:MM:SS
            cant_turnos = int(row['turnos_enviados'])  # Cantidad de turnos para esta hora
            
            # Formatear a HH:MM para mostrar
            try:
                hora_obj = datetime.strptime(hora_turno, '%H:%M:%S').time()
                turno_nombre = hora_obj.strftime('%H:%M')
            except:
                turno_nombre = hora_turno[:5] if len(hora_turno) >= 5 else hora_turno
            
            key = (vehiculo, fecha_str)
            if key not in turnos_por_vehiculo_fecha:
                turnos_por_vehiculo_fecha[key] = {'turnos': [], 'total': 0}
            turnos_por_vehiculo_fecha[key]['turnos'].append(turno_nombre)
            turnos_por_vehiculo_fecha[key]['total'] += cant_turnos
        
        print(f"📊 DEBUG DICT TURNOS - Total vehículos con turnos: {len(turnos_por_vehiculo_fecha)}")
        # Mostrar primeros 5 para debug
        for i, (key, value) in enumerate(list(turnos_por_vehiculo_fecha.items())[:5]):
            print(f"   {key}: {value}")
    
    # Función auxiliar para convertir a date de forma segura
    def to_date(fecha_obj):
        """Convierte datetime o date a date"""
        if isinstance(fecha_obj, datetime):
            return fecha_obj.date()
        return fecha_obj  # Ya es date
    
    # Ya no necesitamos asignar actividades a turnos del mantenedor
    # Ahora trabajamos directamente con las horas de inicio de los turnos enviados
    
    # 5. OBTENER TODAS LAS FECHAS DEL RANGO
    fechas_rango = []
    fecha_actual = fecha_inicio
    while fecha_actual <= fecha_fin:
        fechas_rango.append(fecha_actual)
        fecha_actual += timedelta(days=1)
    
    # 6. COMBINAR INFORMACIÓN - ITERAR POR FECHA Y VEHÍCULO
    datos_completos_lista = []
    
    for fecha_proceso in fechas_rango:
        for _, vehiculo in df_flota.iterrows():
            codigo_tanque = int(vehiculo['codigo_tanque'])
            
            # Obtener turnos enviados con sus horas de inicio PARA ESTA FECHA
            fecha_key = fecha_proceso.strftime('%Y-%m-%d')
            info_turnos = turnos_por_vehiculo_fecha.get((codigo_tanque, fecha_key), {'turnos': [], 'total': 0})
            turnos_vehiculo_fecha = info_turnos['turnos']
            turnos_enviados_total = info_turnos['total']
            
            # DEBUG para primeros 3 vehículos
            if codigo_tanque in [4348, 4395, 8192] and fecha_proceso.day == 15:
                print(f"\n🔍 DEBUG VEHICULO {codigo_tanque} - Fecha: {fecha_key}")
                print(f"   Info turnos: {info_turnos}")
                print(f"   Turnos vehiculo fecha: {turnos_vehiculo_fecha}")
                print(f"   Total turnos enviados: {turnos_enviados_total}")
            
            # Si no tiene turnos enviados en esta fecha
            if not turnos_vehiculo_fecha:
                # Calcular totales sin separación por turno PARA ESTA FECHA
                conexion_rco = 0
                if not df_rco.empty and 'codigo_tanque' in df_rco.columns:
                    rco_rows = df_rco[
                        (df_rco['codigo_tanque'] == codigo_tanque) &
                        (df_rco['fecha'] == to_date(fecha_proceso))
                    ]
                    if not rco_rows.empty:
                        conexion_rco = int(rco_rows['conexiones_rco'].sum())
                
                entregado_totalmente = 0
                en_ruta = 0
                planificado = 0
                
                if not df_entregas.empty:
                    entregas_vehiculo = df_entregas[
                        (df_entregas['vehiclereal'] == codigo_tanque) &
                        (df_entregas['fecha'] == to_date(fecha_proceso))
                    ]
                    
                    if not entregas_vehiculo.empty:
                        entregas_completadas = entregas_vehiculo[
                            entregas_vehiculo['estado'].isin(['Entregado totalmente', 'Recibí Conforme'])
                        ]
                        entregado_totalmente = int(entregas_completadas['cantidad'].sum()) if not entregas_completadas.empty else 0
                        
                        en_ruta_rows = entregas_vehiculo[entregas_vehiculo['estado'] == 'En Ruta']
                        en_ruta = int(en_ruta_rows['cantidad'].sum()) if not en_ruta_rows.empty else 0
                        
                        planificado_rows = entregas_vehiculo[entregas_vehiculo['estado'] == 'Planificado']
                        planificado = int(planificado_rows['cantidad'].sum()) if not planificado_rows.empty else 0
                
                # Solo agregar si hay actividad
                if conexion_rco > 0 or entregado_totalmente > 0 or en_ruta > 0 or planificado > 0:
                    datos_completos_lista.append({
                        'Camion': codigo_tanque,
                        'Fecha': fecha_proceso.strftime('%d-%b'),
                        'fecha_completa': fecha_proceso,
                        'Turno': 'N/A',
                        'Hora_inicio_turno': None,
                        'Entregado_totalmente': entregado_totalmente,
                        'En_ruta': en_ruta,
                        'Planificado': planificado,
                        'Turnos_enviados': 0,
                        'Conexion_RCO': conexion_rco,
                        'Transportista': vehiculo['nombre_transportista'],
                        '¿Es licitado?': 'Si'
                    })
            else:
                # Tiene turnos enviados - crear un registro por turno usando la hora de inicio
                for turno_nombre in turnos_vehiculo_fecha:
                    
                    # Calcular totales para este turno (sin separación por actividad)
                    # Ya que no podemos distinguir qué RCO o entrega pertenece a qué turno
                    # sin conocer la hora del turno, agregamos todo al día
                    conexion_rco = 0
                    if not df_rco.empty and 'codigo_tanque' in df_rco.columns:
                        rco_rows = df_rco[
                            (df_rco['codigo_tanque'] == codigo_tanque) &
                            (df_rco['fecha'] == to_date(fecha_proceso))
                        ]
                        if not rco_rows.empty:
                            conexion_rco = int(rco_rows['conexiones_rco'].sum())
                    
                    entregado_totalmente = 0
                    en_ruta = 0
                    planificado = 0
                    
                    if not df_entregas.empty:
                        entregas_vehiculo = df_entregas[
                            (df_entregas['vehiclereal'] == codigo_tanque) &
                            (df_entregas['fecha'] == to_date(fecha_proceso))
                        ]
                        
                        if not entregas_vehiculo.empty:
                            entregas_completadas = entregas_vehiculo[
                                entregas_vehiculo['estado'].isin(['Entregado totalmente', 'Recibí Conforme'])
                            ]
                            entregado_totalmente = int(entregas_completadas['cantidad'].sum()) if not entregas_completadas.empty else 0
                            
                            en_ruta_rows = entregas_vehiculo[entregas_vehiculo['estado'] == 'En Ruta']
                            en_ruta = int(en_ruta_rows['cantidad'].sum()) if not en_ruta_rows.empty else 0
                            
                            planificado_rows = entregas_vehiculo[entregas_vehiculo['estado'] == 'Planificado']
                            planificado = int(planificado_rows['cantidad'].sum()) if not planificado_rows.empty else 0
                    
                    # Agregar registro para este turno
                    # Solo el primer turno muestra el total de turnos enviados
                    es_primer_turno = (turno_nombre == turnos_vehiculo_fecha[0])
                    
                    datos_completos_lista.append({
                        'Camion': codigo_tanque,
                        'Fecha': fecha_proceso.strftime('%d-%b'),
                        'fecha_completa': fecha_proceso,
                        'Turno': turno_nombre,
                        'Hora_inicio_turno': turno_nombre,
                        'Entregado_totalmente': entregado_totalmente,
                        'En_ruta': en_ruta,
                        'Planificado': planificado,
                        'Turnos_enviados': turnos_enviados_total if es_primer_turno else 0,
                        'Conexion_RCO': conexion_rco,
                        'Transportista': vehiculo['nombre_transportista'],
                        '¿Es licitado?': 'Si'
                    })
    
    if not datos_completos_lista:
        print("⚠️ No hay datos con actividad para esta fecha")
        return None
    
    df_resultado = pd.DataFrame(datos_completos_lista)
    print(f"✅ Datos completos: {len(df_resultado)} registros con actividad")
    
    return df_resultado


@cache.memoize(timeout=900)
def procesar_datos_completos_con_turnos(df_scr, df_turnos, df_rco, fecha_inicio, fecha_fin, minera_nombre=None):
    """
    Procesar y combinar datos de SCR, Turnos y RCO.
    NUEVA VERSIÓN: Granularidad a nivel de [Camión, Fecha, Turno]
    
    Returns:
        DataFrame con estructura:
        [Camion, Fecha, fecha_completa, Turno, Hora_inicio_turno,
         Entregado_totalmente, En_ruta, Planificado, Turnos_enviados,
         Conexion_RCO, Transportista, ¿Es licitado?]
    """
    print(f"🔄 [CACHE MISS] Procesando datos completos CON TURNOS para {minera_nombre}")
    
    # Cargar mantenedor al inicio para usarlo en asignación de turnos
    df_mantenedor = cargar_mantenedor_flota()
    
    # Función para asignar turno basándose en la hora del evento
    def asignar_turno_a_evento(hora_evento, camion, fecha_str, df_mantenedor_local):
        """Asignar el turno correcto basándose en la hora del evento"""
        if pd.isna(hora_evento) or hora_evento is None:
            return 'Sin Turno'
        
        # Obtener configuración de turnos del camión
        config_camion = df_mantenedor_local[df_mantenedor_local['Camion'] == camion]
        if config_camion.empty:
            return 'Sin Turno'
        
        turnos_config = config_camion.iloc[0]['Turnos_parseados']
        if not turnos_config or len(turnos_config) == 0:
            return 'Sin Turno'
        
        # Convertir hora_evento a objeto time si es string
        if isinstance(hora_evento, str):
            try:
                hora_evento = datetime.strptime(hora_evento, '%H:%M:%S').time()
            except:
                return 'Sin Turno'
        
        # Buscar el turno que corresponde a esta hora
        # Asumimos que un evento pertenece al turno si ocurrió después del inicio del turno
        # y antes del inicio del siguiente turno (o fin del día)
        turnos_con_hora = []
        for turno in turnos_config:
            try:
                hora_inicio = datetime.strptime(turno['hora_inicio'], '%H:%M:%S').time()
                turnos_con_hora.append({
                    'nombre': turno['turno_nombre'],
                    'hora_inicio': hora_inicio
                })
            except:
                continue
        
        # Ordenar turnos por hora de inicio
        turnos_con_hora.sort(key=lambda x: x['hora_inicio'])
        
        # Buscar en qué turno cae el evento
        turno_asignado = 'Sin Turno'
        for i, turno in enumerate(turnos_con_hora):
            hora_inicio_turno = turno['hora_inicio']
            
            # Obtener hora fin del turno (inicio del siguiente turno o fin del día)
            if i < len(turnos_con_hora) - 1:
                hora_fin_turno = turnos_con_hora[i + 1]['hora_inicio']
            else:
                # Último turno del día, va hasta el inicio del primer turno del día siguiente
                # Consideramos que va hasta las 23:59:59
                hora_fin_turno = time(23, 59, 59)
            
            # Verificar si el evento cae en este turno
            if hora_inicio_turno <= hora_evento < hora_fin_turno:
                turno_asignado = turno['nombre']
                break
            
            # Caso especial: si es el último turno y el evento es después de medianoche
            # pero antes del primer turno del día
            if i == len(turnos_con_hora) - 1 and hora_evento < turnos_con_hora[0]['hora_inicio']:
                turno_asignado = turno['nombre']
        
        return turno_asignado
    
    # Obtener vehículos únicos del SCR
    vehiculos_totales = df_scr['vehiclereal'].dropna()
    vehiculos_numericos = pd.to_numeric(vehiculos_totales, errors='coerce')
    vehiculos_validos = vehiculos_numericos[vehiculos_numericos > 0].astype(int).unique()
    vehiculos_totales = np.sort(vehiculos_validos)
    
    # Obtener vehículos licitados
    if minera_nombre:
        vehiculos_licitados = obtener_vehiculos_licitados_por_minera(minera_nombre)
    else:
        vehiculos_licitados = [4002, 4003, 4049, 8054, 8120, 8348, 8820]
    
    # Crear tabla base: [Camión x Fecha]
    fechas_rango = pd.date_range(start=fecha_inicio, end=fecha_fin, freq='D')
    indices = list(itertools.product(vehiculos_totales, fechas_rango))
    
    tabla_base = pd.DataFrame(indices, columns=['Camion', 'fecha_completa'])
    tabla_base['Fecha'] = tabla_base['fecha_completa'].dt.strftime('%d-%b')
    tabla_base['fecha_para_merge'] = tabla_base['fecha_completa'].dt.date
    tabla_base['Camion'] = tabla_base['Camion'].astype(int)
    tabla_base['¿Es licitado?'] = np.where(tabla_base['Camion'].isin(vehiculos_licitados), 'Si', 'No')
    
    # NUEVA LÓGICA: Expandir a nivel de turnos
    if not df_mantenedor.empty:
        tabla_base_expandida = expandir_turnos_a_registros(tabla_base, df_mantenedor)
    else:
        print("⚠️ Mantenedor de flota no disponible, trabajando sin turnos")
        tabla_base_expandida = tabla_base.copy()
        tabla_base_expandida['Turno'] = 'Sin Turno'
        tabla_base_expandida['Hora_inicio_turno'] = None
    
    # Procesar turnos enviados
    if not df_turnos.empty:
        estado_final = (
            df_turnos.sort_values(['id_turno_uuid', 'fecha_hora'])
                     .groupby('id_turno_uuid', as_index=False)
                     .tail(1)[['id_turno_uuid', 'estado', 'id_vehiculo', 'fecha_inicio_utc']]
        )
        
        enviados = estado_final[
            (estado_final['estado'] == 'ENVIADO') &
            (estado_final['id_vehiculo'].isin(vehiculos_totales))
        ].copy()
        
        enviados['fecha'] = pd.to_datetime(enviados['fecha_inicio_utc']).dt.date
        conteo_turnos = enviados.groupby(['id_vehiculo', 'fecha']).size().reset_index(name='Turnos_enviados')
        
        # Merge con tabla expandida
        tabla_base_expandida = tabla_base_expandida.merge(
            conteo_turnos,
            left_on=['Camion', 'fecha_para_merge'],
            right_on=['id_vehiculo', 'fecha'],
            how='left'
        )
        tabla_base_expandida['Turnos_enviados'] = tabla_base_expandida['Turnos_enviados'].fillna(0).astype(int)
        tabla_base_expandida = tabla_base_expandida.drop(columns=['fecha', 'id_vehiculo'], errors='ignore')
    else:
        tabla_base_expandida['Turnos_enviados'] = 0
    
    # Procesar conexiones RCO CON ASIGNACIÓN DE TURNO
    if not df_rco.empty:
        df_rco_filtrado = df_rco[df_rco['codigo_tanque'].isin(vehiculos_totales)].copy()
        
        df_rco_filtrado['fecha_conexion_utc'] = pd.to_datetime(df_rco_filtrado['fecha_conexion_utc'])
        df_rco_filtrado['fecha_conexion_santiago'] = (
            df_rco_filtrado['fecha_conexion_utc']
            .dt.tz_localize('UTC')
            .dt.tz_convert('America/Santiago')
        )
        
        hora_corte = pd.Timestamp('23:30').time()
        df_rco_filtrado['fecha_rco'] = df_rco_filtrado['fecha_conexion_santiago'].apply(
            lambda x: (x + pd.Timedelta(days=1)).date()
            if x.time() >= hora_corte
            else x.date()
        )
        
        # Obtener hora de conexión para asignar turno
        df_rco_filtrado['hora_conexion'] = df_rco_filtrado['fecha_conexion_santiago'].dt.time
        
        # Asignar turno a cada conexión RCO
        df_rco_filtrado['turno_asignado'] = df_rco_filtrado.apply(
            lambda row: asignar_turno_a_evento(
                row['hora_conexion'],
                int(row['codigo_tanque']),
                str(row['fecha_rco']),
                df_mantenedor
            ),
            axis=1
        )
        
        # Contar conexiones por [vehiculo, fecha, turno]
        conteos_rco = (df_rco_filtrado
                      .groupby(['codigo_tanque', 'fecha_rco', 'turno_asignado'])
                      .size()
                      .reset_index(name='Conexion_RCO'))
        
        # Merge con tabla expandida (ahora incluye turno)
        tabla_base_expandida = tabla_base_expandida.merge(
            conteos_rco,
            left_on=['Camion', 'fecha_para_merge', 'Turno'],
            right_on=['codigo_tanque', 'fecha_rco', 'turno_asignado'],
            how='left'
        )
        tabla_base_expandida['Conexion_RCO'] = tabla_base_expandida['Conexion_RCO'].fillna(0).astype(int)
        tabla_base_expandida = tabla_base_expandida.drop(columns=['fecha_rco', 'codigo_tanque', 'turno_asignado'], errors='ignore')
    else:
        tabla_base_expandida['Conexion_RCO'] = 0
    
    # Procesar datos SCR para estados CON ASIGNACIÓN DE TURNO
    df_scr['fecha_corregida'] = pd.to_datetime(df_scr['fechallegadaprog']).dt.strftime('%Y-%m-%d')
    
    # Convertir hora de llegada para obtener la hora del evento
    if 'horallegadaprog' in df_scr.columns:
        # horallegadaprog puede venir como string HH:MM:SS o como datetime
        df_scr['hora_evento'] = pd.to_datetime(df_scr['horallegadaprog'], format='%H:%M:%S', errors='coerce').dt.time
    else:
        df_scr['hora_evento'] = None
    
    # Asignar turno a cada registro SCR
    df_scr['vehiclereal_num'] = pd.to_numeric(df_scr['vehiclereal'], errors='coerce').fillna(0).astype(int)
    df_scr['turno_asignado'] = df_scr.apply(
        lambda row: asignar_turno_a_evento(
            row['hora_evento'], 
            row['vehiclereal_num'], 
            row['fecha_corregida'],
            df_mantenedor
        ),
        axis=1
    )
    
    # Contar eventos por [vehiculo, fecha, turno, estado]
    conteos_scr = (df_scr
                   .groupby(['vehiclereal_num', 'fecha_corregida', 'turno_asignado', 'descrstatu'])
                   .size()
                   .reset_index(name='cantidad'))
    
    # Crear pivot con turno incluido
    pivot_scr = conteos_scr.pivot_table(
        index=['vehiclereal_num', 'fecha_corregida', 'turno_asignado'], 
        columns='descrstatu', 
        values='cantidad', 
        fill_value=0
    ).reset_index()
    
    pivot_scr.columns.name = None
    estados_esperados = ['En Ruta', 'Entregado totalmente', 'Planificado']
    for estado in estados_esperados:
        if estado not in pivot_scr.columns:
            pivot_scr[estado] = 0
    
    # Obtener transportista (con turno)
    carriers = (
        df_scr[['vehiclereal_num', 'fecha_corregida', 'turno_asignado', 'carriername1']]
        .dropna(subset=['carriername1'])
        .drop_duplicates(subset=['vehiclereal_num', 'fecha_corregida', 'turno_asignado'])
        .groupby(['vehiclereal_num', 'fecha_corregida', 'turno_asignado'], as_index=False).first()
    )
    
    # Crear mapeo de transportista por camión desde el mantenedor (fallback)
    transportista_por_camion = {}
    if not df_mantenedor.empty:
        for _, row in df_mantenedor.iterrows():
            camion = row['Camion']
            transportista = row.get('Transportista', 'SIN_INFORMACION')
            transportista_por_camion[camion] = transportista
    
    tabla_base_expandida['fecha_merge_scr'] = tabla_base_expandida['fecha_completa'].dt.strftime('%Y-%m-%d')
    
    # IMPORTANTE: Identificar registros con actividad ANTES del merge con SCR
    # para no perderlos en el filtro final
    tiene_turnos_o_rco = (
        (tabla_base_expandida.get('Turnos_enviados', pd.Series([0]*len(tabla_base_expandida))) > 0) |
        (tabla_base_expandida.get('Conexion_RCO', pd.Series([0]*len(tabla_base_expandida))) > 0)
    )
    tabla_base_expandida['_tiene_actividad_previa'] = tiene_turnos_o_rco
    
    # Merge con datos SCR (ahora incluye turno en la join)
    tabla_final = tabla_base_expandida.merge(
        pivot_scr,
        left_on=['Camion', 'fecha_merge_scr', 'Turno'],
        right_on=['vehiclereal_num', 'fecha_corregida', 'turno_asignado'],
        how='left'
    )
    
    for estado in estados_esperados:
        if estado in tabla_final.columns:
            tabla_final[estado] = tabla_final[estado].fillna(0).astype(int)
    
    # Merge con transportistas (ahora incluye turno)
    tabla_final = tabla_final.merge(
        carriers,
        left_on=['Camion', 'fecha_merge_scr', 'Turno'],
        right_on=['vehiclereal_num', 'fecha_corregida', 'turno_asignado'],
        how='left'
    )
    
    tabla_final['Transportista'] = tabla_final.get('carriername1', pd.Series([None]*len(tabla_final)))
    tabla_final['Transportista'] = tabla_final['Transportista'].replace('', pd.NA)
    
    # Prioridad 1: Transportista del SCR por camión (sin considerar turno, ya que puede no tener actividad en ese turno específico)
    per_truck = carriers.groupby('vehiclereal_num', as_index=False)['carriername1'].first().set_index('vehiclereal_num')['carriername1'].to_dict()
    tabla_final['Transportista'] = tabla_final['Transportista'].fillna(tabla_final['Camion'].map(per_truck))
    
    # Prioridad 2: Transportista del mantenedor (fallback para turnos sin actividad)
    tabla_final['Transportista'] = tabla_final['Transportista'].fillna(tabla_final['Camion'].map(transportista_por_camion))
    
    # Último fallback
    tabla_final['Transportista'] = tabla_final['Transportista'].fillna('SIN_INFORMACION').astype(str)
    
    # Limpiar columnas auxiliares
    tabla_final = tabla_final.drop(['fecha_merge_scr', 'vehiclereal_num', 'fecha_corregida', 'turno_asignado', 'carriername1', 'fecha_para_merge'], axis=1, errors='ignore')
    
    # Renombrar columnas
    tabla_final = tabla_final.rename(columns={
        'En Ruta': 'En_ruta',
        'Entregado totalmente': 'Entregado_totalmente',
        'Turnos_enviados': 'Turnos_enviados'
    })
    
    print(f"📊 DEBUG - Registros antes de filtrar: {len(tabla_final)}")
    print(f"📊 DEBUG - Registros con Turnos_enviados > 0: {(tabla_final['Turnos_enviados'] > 0).sum() if 'Turnos_enviados' in tabla_final.columns else 0}")
    print(f"📊 DEBUG - Registros con Conexion_RCO > 0: {(tabla_final['Conexion_RCO'] > 0).sum() if 'Conexion_RCO' in tabla_final.columns else 0}")
    print(f"📊 DEBUG - Registros con Entregado_totalmente > 0: {(tabla_final['Entregado_totalmente'] > 0).sum() if 'Entregado_totalmente' in tabla_final.columns else 0}")
    
    # NO filtrar registros - mantener TODOS los que tienen cualquier actividad
    # Esto incluye: Turnos_enviados > 0, Conexion_RCO > 0, o Entregado_totalmente > 0
    
    # Asegurar que las columnas existan antes de filtrar
    for col in ['Turnos_enviados', 'Conexion_RCO', 'Entregado_totalmente', 'En_ruta', 'Planificado']:
        if col not in tabla_final.columns:
            tabla_final[col] = 0
    
    # Identificar registros con actividad (incluyendo la actividad previa al merge)
    tiene_actividad = (
        (tabla_final['Turnos_enviados'] > 0) |
        (tabla_final['Conexion_RCO'] > 0) |
        (tabla_final['Entregado_totalmente'] > 0) |
        (tabla_final['En_ruta'] > 0) |
        (tabla_final['Planificado'] > 0) |
        tabla_final.get('_tiene_actividad_previa', False)  # Incluir registros que tenían actividad antes del merge
    )
    
    print(f"📊 DEBUG - Registros con actividad (incluyendo previa): {tiene_actividad.sum()}")
    print(f"📊 DEBUG - Registros con SIN_INFORMACION: {(tabla_final['Transportista'] == 'SIN_INFORMACION').sum()}")
    
    # Filtrar solo registros con actividad Y con transportista válido
    tabla_final = tabla_final[
        tiene_actividad & (tabla_final['Transportista'] != 'SIN_INFORMACION')
    ]
    
    # Limpiar columna auxiliar
    tabla_final = tabla_final.drop('_tiene_actividad_previa', axis=1, errors='ignore')
    
    print(f"✅ Registros finales después de filtrar por actividad: {len(tabla_final)}")
    
    # IMPORTANTE: Reorganizar columnas para que Turno esté en posición destacada
    columnas_ordenadas = [
        'Camion', 'Fecha', 'fecha_completa', 'Turno', 'Hora_inicio_turno',
        'Entregado_totalmente', 'En_ruta', 'Planificado', 
        'Turnos_enviados', 'Conexion_RCO',
        'Transportista', '¿Es licitado?'
    ]
    
    # Asegurar que todas las columnas existan
    for col in columnas_ordenadas:
        if col not in tabla_final.columns:
            if col == 'Turno':
                tabla_final[col] = 'Sin Turno'
            elif col == 'Hora_inicio_turno':
                tabla_final[col] = None
            else:
                tabla_final[col] = 0
    
    tabla_final = tabla_final[columnas_ordenadas]
    
    return tabla_final


# ============================================
# FUNCIÓN DE DISPONIBILIDAD (ACTUALIZADA PARA TURNOS)
# ============================================
def calcular_disponibilidad_operacional(datos_completos, banda_total=None):
    """
    Calcular la Disponibilidad Operacional.
    ACTUALIZADO: Considera granularidad por turno.
    
    La disponibilidad se calcula a nivel de [Camión-Fecha-Turno]:
    - Criterio A: Turno con entrega realizada
    - Criterio B: Turno sin entrega PERO con conexión RCO
    
    Args:
        datos_completos: DataFrame con columnas [..., Turno, ...]
        banda_total: Banda total del transportista
    
    Returns:
        dict con métricas de disponibilidad
    """
    if datos_completos is None or datos_completos.empty:
        return {
            'entregas_exitosas_total': 0,
            'entregas_criterio_a': 0,
            'entregas_criterio_b': 0,
            'turnos_enviados_total': 0,
            'disponibilidad_porcentaje': 0.0,
            'entregas_licitadas': 0,
            'entregas_totales': 0,
            'disponibilidad_licitada_porcentaje': 0.0
        }
    
    # Criterio A: Entregado correctamente (por turno)
    entregas_criterio_a = int(datos_completos['Entregado_totalmente'].sum())
    
    # Criterio B: No entregado PERO con conexión RCO (por turno)
    criterio_b_mask = (datos_completos['Entregado_totalmente'] == 0) & (datos_completos['Conexion_RCO'] > 0)
    entregas_criterio_b = int(criterio_b_mask.sum())
    
    entregas_exitosas_total = entregas_criterio_a + entregas_criterio_b
    
    turnos_enviados_total = int(datos_completos['Turnos_enviados'].sum())
    
    denominador = banda_total if banda_total is not None and banda_total > 0 else turnos_enviados_total
    
    if denominador > 0:
        disponibilidad_porcentaje = (entregas_exitosas_total / denominador) * 100
    else:
        disponibilidad_porcentaje = 0.0
    
    entregas_totales = int(datos_completos['Entregado_totalmente'].sum())
    entregas_licitadas = int(
        datos_completos[datos_completos['¿Es licitado?'] == 'Si']['Entregado_totalmente'].sum()
    )
    
    if entregas_totales > 0:
        disponibilidad_licitada_porcentaje = (entregas_licitadas / entregas_totales) * 100
    else:
        disponibilidad_licitada_porcentaje = 0.0
    
    return {
        'entregas_exitosas_total': entregas_exitosas_total,
        'entregas_criterio_a': entregas_criterio_a,
        'entregas_criterio_b': entregas_criterio_b,
        'turnos_enviados_total': turnos_enviados_total,
        'banda_total': denominador,
        'disponibilidad_porcentaje': round(disponibilidad_porcentaje, 1),
        'entregas_licitadas': entregas_licitadas,
        'entregas_totales': entregas_totales,
        'disponibilidad_licitada_porcentaje': round(disponibilidad_licitada_porcentaje, 1)
    }


# ============================================
# FUNCIONES AUXILIARES (COPIAR DEL ORIGINAL)
# ============================================
# NOTA: Las siguientes funciones se mantienen igual que en el original.
# Solo se listan las firmas; debes copiar la implementación completa.

@cache.memoize(timeout=900)
def obtener_datos_desde_athena(minera_nombre, fecha_inicio, fecha_fin):
    """
    Obtener datos desde Athena siguiendo la lógica de athena_test.py
    NOTA: Esta función está cacheada por 15 minutos para reducir queries costosas
    """
    print(f"🔄 [CACHE MISS] Ejecutando query SCR para {minera_nombre} ({fecha_inicio} - {fecha_fin})")
    
    if not USE_ATHENA:
        return None

    # Manejar el caso especial de CODELCO
    if minera_nombre == 'CODELCO':
        mineras_codelco = obtener_mineras_codelco()
        # Crear query para todas las mineras de CODELCO
        conditions = []
        for minera in mineras_codelco:
            minera_safe = minera.replace("'", "''")
            conditions.append(f"vtext LIKE '{minera_safe}'")
        
        # OPTIMIZACIÓN: Seleccionar solo columnas necesarias + hora para asignación de turno
        query = f"""
        SELECT 
            fechallegadaprog,
            vtext,
            vehiclereal,
            carriername1,
            descrstatu,
            horallegadaprog
        FROM {DATABASE_ATHENA}.{TABLA_PEDIDOS}
        WHERE fechallegadaprog >= '{fecha_inicio.strftime('%Y-%m-%d')}'
          AND fechallegadaprog <= '{fecha_fin.strftime('%Y-%m-%d')}'
          AND ({' OR '.join(conditions)})
          AND carriername1 IS NOT NULL
          AND carriername1 != 'SIN_INFORMACION'
        """
    else:
        # Caso normal para mineras individuales
        minera_safe = minera_nombre.replace("'", "''")
        # OPTIMIZACIÓN: Seleccionar solo columnas necesarias + hora para asignación de turno
        query = f"""
        SELECT 
            fechallegadaprog,
            vtext,
            vehiclereal,
            carriername1,
            descrstatu,
            horallegadaprog
        FROM {DATABASE_ATHENA}.{TABLA_PEDIDOS}
        WHERE fechallegadaprog >= '{fecha_inicio.strftime('%Y-%m-%d')}'
          AND fechallegadaprog <= '{fecha_fin.strftime('%Y-%m-%d')}'
          AND vtext LIKE '{minera_safe}'
          AND carriername1 IS NOT NULL
          AND carriername1 != 'SIN_INFORMACION'
        """

    try:
        df = sql_athena(query)
        if df.empty:
            return None

        # Procesar datos según lógica de athena_test.py
        df['Fecha'] = pd.to_datetime(df['fechallegadaprog'], format='%Y-%m-%d', errors='coerce')

        # Calcular entregas (descrstatu) - igual que en athena_test.py
        df['Entregado totalmente'] = (
            (df['descrstatu'] == 'Entregado totalmente') |
            (df['descrstatu'] == 'Recibí Conforme')
        ).astype(int)

        # Limpiar transportista
        df['Transportista'] = df['carriername1'].fillna('SIN_INFORMACION')
        
        # Excluir registros con transportista 'SIN_INFORMACION'
        df = df[df['Transportista'] != 'SIN_INFORMACION']
        
        # Verificar si quedaron datos después del filtro
        if df.empty:
            return None

        return df

    except Exception as e:
        print(f"Error obteniendo datos de Athena: {e}")
        return None


def obtener_vehiculos_licitados_por_minera(minera_nombre):
    """
    Obtener lista de códigos de vehículos licitados para una minera
    
    Args:
        minera_nombre: Nombre de la minera
    
    Returns:
        Lista de códigos de vehículos (int)
    """
    df_flota = obtener_flota_licitada_athena(minera_nombre)
    
    if df_flota.empty:
        # Fallback a lista hardcodeada (mantener compatibilidad)
        return [4002, 4003, 4049, 8054, 8120, 8348, 8820]
    
    return sorted(df_flota['codigo_tanque'].unique().tolist())


@cache.cached(timeout=3600, key_prefix='mineras_athena')  # Cache por 1 hora
def obtener_mineras_athena():
    """Obtener lista de mineras predefinidas"""
    # Lista fija de mineras requeridas
    mineras_predefinidas = [
        'MINA LA ESCONDIDA',
        'QUADRA SIERRA GORDA', 
        'ANDINA',
        'EL TENIENTE',
        'CASERONES',
        'SALARES NORTE',
        'MINERA CANDELARIA',
        'LOS BRONCES',
        'CODELCO'  # Agrupa: MINISTRO HALES, RADOMIRO TOMIC, CHUQUICAMATA, MINA GABY
    ]
    return mineras_predefinidas

@cache.cached(timeout=1800, key_prefix='transportistas_global')  # Cache por 30 minutos
def obtener_transportistas_global():
    """Obtener transportistas únicos de la tabla completa (para administración)."""
    if not USE_ATHENA:
        return []
    # OPTIMIZACIÓN: Agregar límite y filtro de SIN_INFORMACION en query
    query = f"""
    SELECT DISTINCT carriername1 as transportista 
    FROM {DATABASE_ATHENA}.{TABLA_PEDIDOS} 
    WHERE carriername1 IS NOT NULL 
      AND carriername1 != 'SIN_INFORMACION'
    ORDER BY transportista
    """
    try:
        df = sql_athena(query)
        if df.empty:
            return []
        # Filtrar 'SIN_INFORMACION' de la lista
        transportistas = sorted(df['transportista'].dropna().unique().tolist())
        return [t for t in transportistas if t != 'SIN_INFORMACION']
    except Exception as e:
        print(f"Error obteniendo transportistas desde Athena: {e}")
        return []

def obtener_mapeo_uso_mineras():
    """Obtener mapeo entre valores de 'uso' en tabla flota y nombres de mineras"""
    mapeo_uso = {
        # Codelco (agrupación)
        'RADOMIRO TOMIC': 'CODELCO',
        'MINISTRO HALES': 'CODELCO',
        'CODELCO CHUQUI': 'CODELCO',
        'CODELCO CHUQUI PC': 'CODELCO',
        'APOYO CODELCO': 'CODELCO',
        
        # Mineras individuales
        'CASERONES': 'CASERONES',
        'CODELCO TENIENTE': 'EL TENIENTE',
        'LOS BRONCES': 'LOS BRONCES',
        'CODELCO ANDINA': 'ANDINA',
        'QUADRA': 'QUADRA SIERRA GORDA',
        'SALARES NORTE': 'SALARES NORTE',
        'CANDELARIA': 'MINERA CANDELARIA',
        'MEL': 'MINA LA ESCONDIDA'
    }
    return mapeo_uso

@cache.memoize(timeout=600)  # Cache por 10 minutos con parámetros
def obtener_flota_licitada_athena(minera_nombre):
    """
    Obtener vehículos de flota licitada desde Excel local o Athena (fallback)
    
    Prioridad:
    1. Excel local (Flota 20260108.xlsx) - NO se sube a Git
    2. Athena (fallback si no existe el Excel)
    
    Args:
        minera_nombre: Nombre de la minera (según nomenclatura de la app)
    
    Returns:
        DataFrame con columnas: codigo_tanque, nombre_transportista, uso
        o DataFrame vacío si no hay datos o hay error
    """
    # PRIORIDAD 1: Intentar leer desde archivo TXT tabular
    # Buscar en múltiples ubicaciones (Render Secret Files + local)
    rutas_txt_posibles = [
        '/etc/secrets/mantenedor_flota.txt',  # Render Secret Files (ubicación alternativa)
        os.path.join(os.path.dirname(__file__), 'mantenedor_flota.txt'),  # Raíz de la app (Render o local)
        os.path.join(os.path.dirname(__file__), 'Flota_20260108 - copia.txt')  # Nombre alternativo local
    ]
    
    ruta_txt = None
    for ruta in rutas_txt_posibles:
        if os.path.exists(ruta):
            ruta_txt = ruta
            break
    
    if ruta_txt:
        try:
            print(f"\n{'='*80}")
            print(f"📄 LEYENDO FLOTA DESDE TXT LOCAL")
            print(f"   Ruta: {ruta_txt}")
            print(f"   Minera solicitada: {minera_nombre}")
            
            # Intentar múltiples codificaciones
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            df_txt = None
            encoding_used = None
            
            for encoding in encodings:
                try:
                    df_txt = pd.read_csv(ruta_txt, sep='\t', encoding=encoding)
                    encoding_used = encoding
                    print(f"   ✓ TXT cargado exitosamente con codificación: {encoding}")
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            
            if df_txt is None:
                raise Exception(f"No se pudo decodificar el archivo con ninguna codificación: {encodings}")
            
            print(f"   ✓ Total de registros en TXT: {len(df_txt)}")
            print(f"   ✓ Columnas encontradas: {list(df_txt.columns)}")
            
            # Validar columnas requeridas
            columnas_requeridas = ['Equipo', 'Transportista', 'Uso']
            if all(col in df_txt.columns for col in columnas_requeridas):
                # Renombrar columnas al formato esperado por la aplicación
                df_txt = df_txt.rename(columns={
                    'Equipo': 'codigo_tanque',
                    'Transportista': 'nombre_transportista',
                    'Uso': 'uso'
                })
                
                # Obtener mapeo de uso a mineras
                mapeo_uso = obtener_mapeo_uso_mineras()
                
                # Invertir el mapeo para buscar por minera
                uso_values = [uso for uso, minera in mapeo_uso.items() if minera == minera_nombre]
                print(f"   ✓ Valores de 'Uso' buscados para {minera_nombre}: {uso_values}")
                
                if not uso_values:
                    print(f"   ⚠️ No se encontró mapeo de uso para minera: {minera_nombre}")
                    print(f"{'='*80}\n")
                    return pd.DataFrame()
                
                # Mostrar valores únicos de 'uso' en el TXT para debug
                print(f"   ✓ Valores únicos de 'Uso' en TXT: {df_txt['uso'].unique().tolist()}")
                
                # Filtrar por minera (todos están activos, no hay columna 'estado')
                df_flota = df_txt[df_txt['uso'].isin(uso_values)].copy()
                print(f"   ✓ Vehículos encontrados después de filtrar por uso: {len(df_flota)}")
                
                if df_flota.empty:
                    print(f"   ⚠️ No se encontraron vehículos para {minera_nombre} en TXT")
                    print(f"{'='*80}\n")
                    return pd.DataFrame()
                
                # Convertir codigo_tanque a int
                df_flota['codigo_tanque'] = df_flota['codigo_tanque'].astype(int)
                print(f"   ✓ Códigos de tanque convertidos a enteros")
                
                # Eliminar duplicados
                df_flota = df_flota.drop_duplicates(subset=['codigo_tanque'], keep='first')
                print(f"   ✓ Duplicados eliminados (si existían)")
                
                # Mostrar primeros vehículos como muestra
                print(f"   ✓ Muestra de vehículos cargados:")
                for idx, row in df_flota.head(3).iterrows():
                    print(f"      - Equipo {row['codigo_tanque']}: {row['nombre_transportista']} (Uso: {row['uso']})")
                
                print(f"   ✅ FLOTA CARGADA EXITOSAMENTE DESDE TXT")
                print(f"   ✅ Total de vehículos: {len(df_flota)} para {minera_nombre}")
                print(f"{'='*80}\n")
                return df_flota
            else:
                print(f"   ❌ TXT no tiene las columnas requeridas: {columnas_requeridas}")
                print(f"   ❌ Columnas encontradas: {list(df_txt.columns)}")
                print(f"{'='*80}\n")
        except Exception as e:
            print(f"   ❌ Error leyendo TXT local: {e}")
            print(f"   ↪️  Intentando con Excel como fallback...")
            print(f"{'='*80}\n")
    else:
        print(f"\n{'='*80}")
        print(f"📁 TXT NO ENCONTRADO en ninguna ubicación")
        print(f"   Ubicaciones buscadas:")
        for ruta in rutas_txt_posibles:
            print(f"   - {ruta}")
        print(f"↪️  Intentando con Excel como fallback...")
        print(f"{'='*80}\n")
    
    # PRIORIDAD 2: Intentar leer desde Excel local
    ruta_excel = os.path.join(os.path.dirname(__file__), 'Flota_20260108.xlsx')
    
    if os.path.exists(ruta_excel):
        try:
            print(f"\n{'='*80}")
            print(f"📊 LEYENDO FLOTA DESDE EXCEL LOCAL")
            print(f"   Ruta: {ruta_excel}")
            print(f"   Minera solicitada: {minera_nombre}")
            df_excel = pd.read_excel(ruta_excel)
            print(f"   ✓ Excel cargado exitosamente")
            print(f"   ✓ Total de registros en Excel: {len(df_excel)}")
            print(f"   ✓ Columnas encontradas: {list(df_excel.columns)}")
            
            # Validar columnas requeridas (según estructura del nuevo Excel)
            columnas_requeridas = ['Equipo', 'Transportista', 'Uso']
            if all(col in df_excel.columns for col in columnas_requeridas):
                # Renombrar columnas al formato esperado por la aplicación
                df_excel = df_excel.rename(columns={
                    'Equipo': 'codigo_tanque',
                    'Transportista': 'nombre_transportista',
                    'Uso': 'uso'
                })
                
                # Obtener mapeo de uso a mineras
                mapeo_uso = obtener_mapeo_uso_mineras()
                
                # Invertir el mapeo para buscar por minera
                uso_values = [uso for uso, minera in mapeo_uso.items() if minera == minera_nombre]
                print(f"   ✓ Valores de 'Uso' buscados para {minera_nombre}: {uso_values}")
                
                if not uso_values:
                    print(f"   ⚠️ No se encontró mapeo de uso para minera: {minera_nombre}")
                    print(f"{'='*80}\n")
                    return pd.DataFrame()
                
                # Mostrar valores únicos de 'uso' en el Excel para debug
                print(f"   ✓ Valores únicos de 'Uso' en Excel: {df_excel['uso'].unique().tolist()}")
                
                # Filtrar por minera (todos están activos, no hay columna 'estado')
                df_flota = df_excel[df_excel['uso'].isin(uso_values)].copy()
                print(f"   ✓ Vehículos encontrados después de filtrar por uso: {len(df_flota)}")
                
                if df_flota.empty:
                    print(f"   ⚠️ No se encontraron vehículos para {minera_nombre} en Excel")
                    print(f"{'='*80}\n")
                    return pd.DataFrame()
                
                # Convertir codigo_tanque a int
                df_flota['codigo_tanque'] = df_flota['codigo_tanque'].astype(int)
                print(f"   ✓ Códigos de tanque convertidos a enteros")
                
                # Eliminar duplicados
                df_flota = df_flota.drop_duplicates(subset=['codigo_tanque'], keep='first')
                print(f"   ✓ Duplicados eliminados (si existían)")
                
                # Mostrar primeros vehículos como muestra
                print(f"   ✓ Muestra de vehículos cargados:")
                for idx, row in df_flota.head(3).iterrows():
                    print(f"      - Equipo {row['codigo_tanque']}: {row['nombre_transportista']} (Uso: {row['uso']})")
                
                print(f"   ✅ FLOTA CARGADA EXITOSAMENTE DESDE EXCEL")
                print(f"   ✅ Total de vehículos: {len(df_flota)} para {minera_nombre}")
                print(f"{'='*80}\n")
                return df_flota
            else:
                print(f"   ❌ Excel no tiene las columnas requeridas: {columnas_requeridas}")
                print(f"   ❌ Columnas encontradas: {list(df_excel.columns)}")
                print(f"{'='*80}\n")
        except Exception as e:
            print(f"   ❌ Error leyendo Excel local: {e}")
            print(f"   ↪️  Intentando con Athena como fallback...")
            print(f"{'='*80}\n")
    else:
        print(f"\n{'='*80}")
        print(f"📁 Excel local NO ENCONTRADO: {ruta_excel}")
        print(f"↪️  Usando Athena como fallback...")
        print(f"{'='*80}\n")
    
    # PRIORIDAD 3: Fallback a Athena si no hay TXT/Excel o hay error
    if not USE_ATHENA:
        return pd.DataFrame()
    
    # Validar que existe la tabla de flota configurada
    if not TABLA_FLOTA:
        print("TABLA_FLOTA no está configurada en las variables de entorno")
        return pd.DataFrame()
    
    # Obtener mapeo de uso a mineras
    mapeo_uso = obtener_mapeo_uso_mineras()
    
    # Invertir el mapeo para buscar por minera
    uso_values = [uso for uso, minera in mapeo_uso.items() if minera == minera_nombre]
    
    if not uso_values:
        print(f"No se encontró mapeo de uso para minera: {minera_nombre}")
        return pd.DataFrame()
    
    # Crear condiciones para la query con escape correcto
    uso_conditions = [f"'{uso}'" for uso in uso_values]
    uso_in_clause = ', '.join(uso_conditions)
    
    query = f"""
    SELECT 
        codigo_tanque,
        nombre_transportista,
        estado,
        uso
    FROM {DATABASE_DISPOMATE}.{TABLA_FLOTA}
    WHERE uso IN ({uso_in_clause})
      AND estado = 'Activo'
    ORDER BY codigo_tanque
    """
    
    try:
        df_flota = sql_athena(query)
        if df_flota.empty:
            print(f"No se encontraron vehículos licitados para {minera_nombre} en Athena")
            return pd.DataFrame()
        
        # Convertir codigo_tanque a int
        df_flota['codigo_tanque'] = df_flota['codigo_tanque'].astype(int)
        
        # Eliminar duplicados - mantener solo un registro por vehículo
        # Priorizar el primer registro encontrado
        df_flota = df_flota.drop_duplicates(subset=['codigo_tanque'], keep='first')
        
        print(f"✅ Flota cargada desde Athena: {len(df_flota)} vehículos para {minera_nombre}")
        return df_flota
    except Exception as e:
        print(f"Error obteniendo flota licitada de Athena: {e}")
        return pd.DataFrame()

@cache.cached(timeout=3600, key_prefix='config_bandas')  # Cache por 1 hora
def obtener_configuracion_bandas_transportista():
    """Obtener configuración de bandas de capacidad por transportista"""
    # Banda estándar por transportista (viajes por día)
    bandas_transportista = {
        # Configuraciones específicas por transportista y minera
        'SOC. DE TRANSP. ILZAUSPE LTDA.': {
            'MINA LA ESCONDIDA': 13,
            'QUADRA SIERRA GORDA': 11,
            'ANDINA': 6,
            'LOS BRONCES': 6
        },
        'TRANSPORTES DE COMBUSTIBLES CHILE L': {
            'MINA LA ESCONDIDA': 14,
            'SALARES NORTE': 4,
            'MINERA CANDELARIA': 9
        },
        'TRANSPORTES SOLUCIONES LOGISTICAS': {
            'MINA LA ESCONDIDA': 11,
            'CODELCO': 11
        },
        'SCP SOTRASER S.A.': {
            'CODELCO': 12
        },
        'TRANSPORTES VIGAL S.A.': {
            'EL TENIENTE': 3,
            'CASERONES': 5,
            'CODELCO': 12
        },
        'SOCIEDAD DE TRANSPORTE NAZAR LTDA': {
            'LOS BRONCES': 6
        }
    }
    return bandas_transportista


def obtener_transportistas_autorizados(minera):
    """Obtener transportistas autorizados para una minera específica"""
    bandas = obtener_configuracion_bandas_transportista()
    transportistas_autorizados = set()
    
    for transportista, mineras_config in bandas.items():
        if minera in mineras_config:
            transportistas_autorizados.add(transportista)
    
    return transportistas_autorizados


def es_transportista_autorizado(transportista, minera):
    """Verificar si un transportista está autorizado para una minera usando palabras clave"""
    transportistas_autorizados = obtener_transportistas_autorizados(minera)
    
    # Normalizar nombre del transportista para búsqueda
    transportista_normalizado = transportista.upper().strip()
    
    # Palabras clave para identificar transportistas (ignorando variaciones de "SOCIEDAD", "DE", "TRANSPORTES", etc.)
    palabras_clave_transportistas = {
        'ILZAUSPE': ['ILZAUSPE'],
        'NAZAR': ['NAZAR'],
        'COMBUSTIBLES CHILE': ['COMBUSTIBLES', 'CHILE'],
        'SOLUCIONES LOGISTICAS': ['SOLUCIONES', 'LOGISTICAS'],
        'SOTRASER': ['SOTRASER'],
        'VIGAL': ['VIGAL'],
    }
    
    # Para cada transportista autorizado, extraer palabras clave
    for nombre_autorizado in transportistas_autorizados:
        nombre_autorizado_upper = nombre_autorizado.upper()
        
        # Buscar si alguna palabra clave del transportista autorizado está en el nombre del DB
        for clave, palabras in palabras_clave_transportistas.items():
            # Si el nombre autorizado contiene la clave
            if any(palabra in nombre_autorizado_upper for palabra in palabras):
                # Y el nombre del DB también contiene todas las palabras de la clave
                if all(palabra in transportista_normalizado for palabra in palabras):
                    return True
    
    # Fallback: buscar exacto o parcial (compatibilidad con casos no cubiertos por palabras clave)
    if transportista_normalizado in [t.upper() for t in transportistas_autorizados]:
        return True
    
    for nombre_autorizado in transportistas_autorizados:
        if (nombre_autorizado.upper() in transportista_normalizado or 
            transportista_normalizado in nombre_autorizado.upper()):
            return True
    
    return False



def calcular_banda_transportista(transportista, minera, bandas_config=None):
    """Calcular la banda (capacidad diaria) de un transportista para una minera específica"""
    if bandas_config is None:
        bandas_config = obtener_configuracion_bandas_transportista()
    
    # Verificar si es transportista autorizado
    if not es_transportista_autorizado(transportista, minera):
        return None  # No tiene banda, se medirá por entregas
    
    # Normalizar nombre del transportista para búsqueda
    transportista_normalizado = transportista.upper().strip()
    
    # Palabras clave para identificar transportistas
    palabras_clave_transportistas = {
        'ILZAUSPE': ['ILZAUSPE'],
        'NAZAR': ['NAZAR'],
        'COMBUSTIBLES CHILE': ['COMBUSTIBLES', 'CHILE'],
        'SOLUCIONES LOGISTICAS': ['SOLUCIONES', 'LOGISTICAS'],
        'SOTRASER': ['SOTRASER'],
        'VIGAL': ['VIGAL'],
    }
    
    # Buscar configuración específica del transportista usando palabras clave
    for nombre_config, mineras_config in bandas_config.items():
        if minera in mineras_config:
            nombre_config_upper = nombre_config.upper()
            
            # Buscar por palabras clave
            for clave, palabras in palabras_clave_transportistas.items():
                # Si el nombre configurado contiene la clave
                if any(palabra in nombre_config_upper for palabra in palabras):
                    # Y el nombre del DB también contiene todas las palabras de la clave
                    if all(palabra in transportista_normalizado for palabra in palabras):
                        return mineras_config[minera]
            
            # Fallback: match exacto o parcial
            if (transportista_normalizado == nombre_config_upper or
                nombre_config_upper in transportista_normalizado or 
                transportista_normalizado in nombre_config_upper):
                return mineras_config[minera]
    
    # Si llegamos aquí, es un error en la configuración
    return None


@cache.cached(timeout=3600, key_prefix='config_viajes')  # Cache por 1 hora
def obtener_configuracion_viajes():
    """Obtener configuración de bandas mínimas y máximas de viajes por minera"""
    configuracion_viajes = {
        'MINA LA ESCONDIDA': {'minimo': 38, 'maximo': 38},
        'QUADRA SIERRA GORDA': {'minimo': 11, 'maximo': 13}, 
        'ANDINA': {'minimo': 6, 'maximo': 7},
        'EL TENIENTE': {'minimo': 3, 'maximo': 4},
        'CASERONES': {'minimo': 5, 'maximo': 6},
        'SALARES NORTE': {'minimo': 4, 'maximo': 5},
        'MINERA CANDELARIA': {'minimo': 9, 'maximo': 9},
        'LOS BRONCES': {'minimo': 12, 'maximo': 12},
        'CODELCO': {'minimo': 36, 'maximo': 42}  
    }
    return configuracion_viajes

@cache.cached(timeout=3600, key_prefix='mineras_codelco')  # Cache por 1 hora
def obtener_mineras_codelco():
    """Obtener lista de mineras que forman parte de CODELCO"""
    return ['MINISTRO HALES', 'RADOMIRO TOMIC', 'CHUQUICAMATA', 'MINA GABY']

def obtener_datos_grafico_athena(minera_nombre, fecha_inicio, fecha_fin, viajes_min, viajes_max):
    """
    Obtener datos para el gráfico de disponibilidad por transportista.
    
    REFACTORIZADA: Usa datos cacheados en lugar de queries directas.
    
    Args:
        minera_nombre: Nombre de la minera
        fecha_inicio: Fecha de inicio (datetime.date o datetime)
        fecha_fin: Fecha de fin (datetime.date o datetime)
        viajes_min: Umbral mínimo de viajes
        viajes_max: Umbral máximo de viajes
    
    Returns:
        dict: Diccionario con estructura:
        {
            'datos': [
                {
                    'fecha': 'DD-MM',
                    'TransportistaA': X,
                    'TransportistaB': Y,
                    'total': Z,
                    'cumple': True/False
                },
                ...
            ],
            'transportistas': ['TransportistaA', 'TransportistaB', ...],
            'minimo': viajes_min,
            'maximo': viajes_max,
            'tipo': 'apilado'
        }
    """
    print(f"📊 [GRAFICO] Obteniendo datos para {minera_nombre} ({fecha_inicio} - {fecha_fin})")
    
    # OPTIMIZACIÓN: Usar función cacheada en lugar de query directa
    datos_completos = obtener_datos_completos_athena(minera_nombre, fecha_inicio, fecha_fin)
    
    if datos_completos is None or datos_completos.empty:
        return {
            'datos': [],
            'minimo': viajes_min,
            'maximo': viajes_max,
            'tipo': 'apilado',
            'transportistas': []
        }
    
    # Agrupar por Transportista, Camión y Fecha (sumar turnos del mismo día)
    datos_agregados = datos_completos.groupby(['Transportista', 'Camion', 'fecha_completa']).agg({
        'Entregado_totalmente': 'sum',  # Suma de entregas de todos los turnos
    }).reset_index()
    
    # Separar transportistas autorizados y otros
    transportistas_unicos = datos_agregados['Transportista'].unique()
    print(f"🚛 DEBUG - Transportistas únicos encontrados: {transportistas_unicos.tolist()}")
    
    transportistas_autorizados_config = obtener_transportistas_autorizados(minera_nombre)
    print(f"🚛 DEBUG - Transportistas autorizados configurados: {transportistas_autorizados_config}")
    
    transportistas_autorizados = []
    otros_transportistas = []
    
    for transportista in transportistas_unicos:
        es_autorizado = es_transportista_autorizado(transportista, minera_nombre)
        print(f"🚛 DEBUG - '{transportista}' autorizado: {es_autorizado}")
        
        if es_autorizado:
            transportistas_autorizados.append(transportista)
        else:
            otros_transportistas.append(transportista)
    
    # Lista final para el gráfico
    transportistas_finales = sorted(transportistas_autorizados)
    if otros_transportistas:
        transportistas_finales.append('OTRO TRANSPORTISTA')
    
    # Crear estructura para gráfico apilado
    fechas_ordenadas = sorted(datos_agregados['fecha_completa'].unique())
    
    datos_apilados = []
    for fecha in fechas_ordenadas:
        fecha_str = fecha.strftime('%d-%m')
        datos_fecha = {'fecha': fecha_str, 'total': 0}
        
        # Agregar datos por transportista autorizado
        for transportista in transportistas_autorizados:
            datos_transportista = datos_agregados[
                (datos_agregados['fecha_completa'] == fecha) & 
                (datos_agregados['Transportista'] == transportista)
            ]
            cantidad = int(datos_transportista['Entregado_totalmente'].sum()) if not datos_transportista.empty else 0
            datos_fecha[transportista] = cantidad
            datos_fecha['total'] += cantidad
        
        # Agregar datos agrupados de "otros" transportistas
        if otros_transportistas:
            cantidad_otros = 0
            for transportista in otros_transportistas:
                datos_transportista = datos_agregados[
                    (datos_agregados['fecha_completa'] == fecha) & 
                    (datos_agregados['Transportista'] == transportista)
                ]
                cantidad_otros += int(datos_transportista['Entregado_totalmente'].sum()) if not datos_transportista.empty else 0
            
            datos_fecha['OTRO TRANSPORTISTA'] = cantidad_otros
            datos_fecha['total'] += cantidad_otros
        
        # Verificar si cumple con el rango
        datos_fecha['cumple'] = viajes_min <= datos_fecha['total'] <= viajes_max
        datos_apilados.append(datos_fecha)
    
    grafico_data = {
        'datos': datos_apilados,
        'minimo': viajes_min,
        'maximo': viajes_max,
        'tipo': 'apilado',
        'transportistas': transportistas_finales
    }
    
    print(f"✅ [GRAFICO] Datos procesados: {len(datos_apilados)} días, {len(transportistas_finales)} transportistas")
    
    return grafico_data

def obtener_datos_matriz_athena(minera_nombre, fecha_inicio, fecha_fin, viajes_min, viajes_max):
    """
    Obtener datos para la matriz de disponibilidad (transportistas x fechas).
    
    REFACTORIZADA: Usa datos cacheados en lugar de queries directas.
    
    Args:
        minera_nombre: Nombre de la minera
        fecha_inicio: Fecha de inicio (datetime.date o datetime)
        fecha_fin: Fecha de fin (datetime.date o datetime)
        viajes_min: Umbral mínimo de viajes
        viajes_max: Umbral máximo de viajes
    
    Returns:
        dict: Diccionario con estructura:
        {
            'transportistas': ['TransportistaA', 'TransportistaB', ...],
            'fechas': ['DD-MM', 'DD-MM', ...],
            'datos': {
                'TransportistaA': {
                    'DD-MM': {
                        'porcentaje': X,
                        'total': Y,
                        'banda': Z,
                        'fecha_full': 'YYYY-MM-DD',
                        'disponibilidad_operacional': {...},
                        'entregas_licitadas': N,
                        'entregas_totales': M
                    }
                }
            },
            'bandas_transportistas': {...}
        }
    """
    print(f"🔲 [MATRIZ] Obteniendo datos para {minera_nombre} ({fecha_inicio} - {fecha_fin})")
    
    # OPTIMIZACIÓN: Usar función cacheada
    datos_completos = obtener_datos_completos_athena(minera_nombre, fecha_inicio, fecha_fin)
    
    if datos_completos is None or datos_completos.empty:
        return {
            'transportistas': [],
            'fechas': [],
            'datos': {},
            'bandas_transportistas': {}
        }
    
    # Agrupar datos por Transportista, Camión y Fecha (sumar turnos del mismo día)
    datos_agregados = datos_completos.groupby(['Transportista', 'Camion', 'fecha_completa']).agg({
        'Entregado_totalmente': 'sum',
        'Conexion_RCO': 'sum',
        'Turnos_enviados': 'sum'
    }).reset_index()
    
    # Separar transportistas autorizados y otros
    transportistas_unicos = datos_agregados['Transportista'].unique()
    transportistas_autorizados = []
    otros_transportistas_data = []
    
    for transportista in transportistas_unicos:
        if es_transportista_autorizado(transportista, minera_nombre):
            transportistas_autorizados.append(transportista)
        else:
            otros_data = datos_agregados[datos_agregados['Transportista'] == transportista]
            otros_transportistas_data.append(otros_data)
    
    # Lista final de transportistas para mostrar
    transportistas_finales = sorted(transportistas_autorizados)
    if otros_transportistas_data:
        transportistas_finales.append('OTRO TRANSPORTISTA')
    
    # Obtener fechas únicas y ordenadas
    fechas_completas = sorted(datos_agregados['fecha_completa'].unique())
    fechas = [d.strftime('%d-%m') for d in fechas_completas]
    
    matriz_data = {
        'transportistas': transportistas_finales,
        'fechas': fechas,
        'datos': {},
        'bandas_transportistas': {}
    }
    
    # Procesar transportistas autorizados
    for transportista_nombre in transportistas_autorizados:
        banda_transportista = calcular_banda_transportista(transportista_nombre, minera_nombre)
        matriz_data['bandas_transportistas'][transportista_nombre] = banda_transportista
        matriz_data['datos'][transportista_nombre] = {}
        
        for fecha_completa in fechas_completas:
            fecha_str = fecha_completa.strftime('%d-%m')
            fecha_full = fecha_completa.strftime('%Y-%m-%d')
            
            # Obtener entregas del transportista en esa fecha
            datos_trans = datos_agregados[
                (datos_agregados['Transportista'] == transportista_nombre) & 
                (datos_agregados['fecha_completa'] == fecha_completa)
            ]
            viajes_realizados = int(datos_trans['Entregado_totalmente'].sum())
            
            # Calcular disponibilidad por banda
            if banda_transportista and banda_transportista > 0:
                disponibilidad_banda = (viajes_realizados / banda_transportista) * 100
            else:
                disponibilidad_banda = 0
            
            # Calcular disponibilidad operacional
            disp_op_data = {'porcentaje': 0, 'turnos_enviados': 0, 'entregas_exitosas': 0}
            entregas_licitadas = 0
            entregas_totales = 0
            
            # Filtrar datos completos por transportista y fecha para cálculo detallado
            datos_dia = datos_completos[
                (datos_completos['Transportista'] == transportista_nombre) &
                (datos_completos['fecha_completa'] == fecha_completa)
            ]
            
            if not datos_dia.empty:
                # NO filtrar - usar TODOS los datos del día para calcular disponibilidad operacional
                # La función calcular_disponibilidad_operacional ya maneja la lógica correctamente
                resultado_disp = calcular_disponibilidad_operacional(datos_dia, banda_transportista)
                disp_op_data = {
                    'porcentaje': resultado_disp['disponibilidad_porcentaje'],
                    'turnos_enviados': resultado_disp['turnos_enviados_total'],
                    'entregas_exitosas': resultado_disp['entregas_exitosas_total']
                }
                entregas_licitadas = resultado_disp['entregas_licitadas']
                entregas_totales = resultado_disp['entregas_totales']
            
            matriz_data['datos'][transportista_nombre][fecha_str] = {
                'porcentaje': round(disp_op_data['porcentaje'], 1),  # Mostrar disponibilidad operacional como principal
                'disponibilidad_banda': round(disponibilidad_banda, 1),  # Agregar disponibilidad por banda separada
                'total': viajes_realizados,
                'banda': banda_transportista,
                'cumplidos': viajes_realizados,
                'transportista_id': 0,
                'fecha_full': fecha_full,
                'disponibilidad_operacional': disp_op_data,
                'entregas_licitadas': entregas_licitadas,
                'entregas_totales': entregas_totales
            }
    
    # Procesar "OTRO TRANSPORTISTA"
    if otros_transportistas_data:
        matriz_data['bandas_transportistas']['OTRO TRANSPORTISTA'] = None
        matriz_data['datos']['OTRO TRANSPORTISTA'] = {}
        
        # Combinar todos los datos de otros transportistas
        otros_combined = pd.concat(otros_transportistas_data, ignore_index=True)
        
        for fecha_completa in fechas_completas:
            fecha_str = fecha_completa.strftime('%d-%m')
            fecha_full = fecha_completa.strftime('%Y-%m-%d')
            
            datos_otros = otros_combined[otros_combined['fecha_completa'] == fecha_completa]
            entregas_realizadas = int(datos_otros['Entregado_totalmente'].sum())
            
            # Calcular disponibilidad operacional para OTROS
            disp_op_otros = {'porcentaje': 0, 'turnos_enviados': 0, 'entregas_exitosas': 0}
            
            # Filtrar datos completos de otros transportistas para esta fecha
            datos_dia_otros = datos_completos[
                (~datos_completos['Transportista'].isin(transportistas_autorizados)) &
                (datos_completos['fecha_completa'] == fecha_completa)
            ]
            
            if not datos_dia_otros.empty:
                # NO filtrar - usar TODOS los datos para el cálculo
                resultado_disp = calcular_disponibilidad_operacional(datos_dia_otros, None)
                disp_op_otros = {
                    'porcentaje': resultado_disp['disponibilidad_porcentaje'],
                    'turnos_enviados': resultado_disp['turnos_enviados_total'],
                    'entregas_exitosas': resultado_disp['entregas_exitosas_total']
                }
            
            matriz_data['datos']['OTRO TRANSPORTISTA'][fecha_str] = {
                'porcentaje': round(disp_op_otros['porcentaje'], 1),  # Usar disponibilidad operacional
                'total': entregas_realizadas,
                'banda': None,
                'cumplidos': entregas_realizadas,
                'transportista_id': 0,
                'fecha_full': fecha_full,
                'es_otro': True,
                'disponibilidad_operacional': disp_op_otros,
                'entregas_licitadas': 0,
                'entregas_totales': entregas_realizadas
            }
    
    print(f"✅ [MATRIZ] Datos procesados: {len(transportistas_finales)} transportistas x {len(fechas)} fechas")
    
    return matriz_data

def exportar_xlsx_matricial(df_export, fecha_inicio, fecha_fin, output_buffer, minera_nombre):
    """
    Generar archivo XLSX con formato matricial por transportista LICITADO
    Solo incluye transportistas que tienen vehículos licitados
    
    Args:
        df_export: DataFrame con los datos
        fecha_inicio: datetime - Fecha de inicio
        fecha_fin: datetime - Fecha de fin
        output_buffer: BytesIO buffer donde guardar el archivo
        minera_nombre: str - Nombre de la minera para obtener bandas
    """
    
    # Crear workbook
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Remover hoja por defecto
    
    # Filtrar solo transportistas LICITADOS
    df_licitados = df_export[df_export['Es_Licitado'] == 'Si'].copy()
    
    if df_licitados.empty:
        # Crear hoja con mensaje
        ws = workbook.create_sheet(title="Sin Datos")
        ws['A1'] = "No se encontraron transportistas con flota licitada para esta minera"
        workbook.save(output_buffer)
        return
    
    # Obtener solo transportistas que tienen vehículos licitados
    transportistas_licitados = sorted(df_licitados['Transportista'].unique())
    
    # Separar datos de transportistas NO licitados para hoja separada
    df_no_licitados = df_export[df_export['Es_Licitado'] == 'No'].copy()
    
    # Obtener rango de fechas
    fechas = pd.date_range(start=fecha_inicio, end=fecha_fin, freq='D')
    num_dias = len(fechas)
    
    # Estilos
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Crear una hoja por transportista LICITADO
    for transportista in transportistas_licitados:
        # Filtrar SOLO datos de flota licitada del transportista
        df_transp = df_licitados[df_licitados['Transportista'] == transportista].copy()
        
        # DEBUG: Ver datos antes de procesar
        print(f"\n📊 DEBUG EXCEL - Transportista: {transportista}")
        print(f"Total registros: {len(df_transp)}")
        if not df_transp.empty:
            print("Columnas:", df_transp.columns.tolist())
            print("Primeros registros:")
            print(df_transp[['Camion', 'Fecha', 'Turno', 'Entregado_totalmente', 'Conexion_RCO']].head(10))
        
        # Crear worksheet
        ws = workbook.create_sheet(title=transportista[:31])  # Límite de 31 caracteres
        
        # SECCIÓN DE CONFIGURACIÓN
        # Fila 1 - Encabezados
        cell_config = ws['C1']
        cell_config.value = 'Configuración de banda minima'
        cell_config.font = Font(bold=True)
        ws.merge_cells('C1:H1')
        
        cell_max = ws['I1']
        cell_max.value = 'Configuración de banda maxima'
        cell_max.font = Font(bold=True)
        ws.merge_cells('I1:P1')
              
        # Fila 2 - Valores de configuración (banda del transportista)
        banda_transportista = calcular_banda_transportista(transportista, minera_nombre)
        banda_valor = banda_transportista if banda_transportista else 0
        
        # Asignar valores ANTES de fusionar las celdas
        ws['C2'] = banda_valor
        ws.merge_cells('C2:H2')
        
        ws['I2'] = banda_valor
        ws.merge_cells('I2:P2')
        
        # Columna de Disponibilidad del periodo
        col_disp_start = 22 + (num_dias * 3)
        ws.cell(row=1, column=col_disp_start).value = 'Disponibilidad del periodo'
        ws.cell(row=1, column=col_disp_start).font = Font(bold=True)
        
        # Calcular disponibilidad del transportista
        resultado_disp = calcular_disponibilidad_operacional(df_transp)
        ws.cell(row=2, column=col_disp_start).value = resultado_disp['disponibilidad_porcentaje'] / 100
        ws.cell(row=2, column=col_disp_start).number_format = '0.00'
        
        # ENCABEZADOS DE DÍAS
        row_dia = 5
        row_criterio = 6
        
        ws['A5'] = 'Día'
        ws['A6'] = 'Camion \\ Criterio'
        ws['B5'] = ''  
        ws['B6'] = 'Turno'
        
        # Aplicar formato a columna Turno
        ws.cell(row=row_criterio, column=2).font = header_font
        ws.cell(row=row_criterio, column=2).fill = header_fill
        ws.cell(row=row_criterio, column=2).alignment = center_alignment
        
        # Configurar encabezados de días y criterios
        current_col = 3  # Columna C (después de Camion y Turno)
        for fecha in fechas:
            dia_numero = fecha.day
            
            # Día (se repite en 3 columnas)
            ws.merge_cells(start_row=row_dia, start_column=current_col, 
                          end_row=row_dia, end_column=current_col + 2)
            cell_dia = ws.cell(row=row_dia, column=current_col)
            cell_dia.value = dia_numero
            cell_dia.alignment = center_alignment
            
            # Criterios 1, 2, 3
            for criterio in [1, 2, 3]:
                ws.cell(row=row_criterio, column=current_col).value = criterio
                ws.cell(row=row_criterio, column=current_col).alignment = center_alignment
                current_col += 1
        
        # Aplicar formato a encabezados
        for col in range(1, current_col):
            for row in [row_dia, row_criterio]:
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
        
        # DATOS DE CAMIONES - AHORA DESAGREGADO POR TURNO
        # Crear una lista de (camion, turno) para iterar
        camiones_turnos = []
        df_transp_sorted = df_transp.sort_values(['Camion', 'Turno'])
        
        for _, row in df_transp_sorted.iterrows():
            camion = row['Camion']
            turno = row['Turno']
            # Solo agregar combinaciones únicas de camion-turno
            if (camion, turno) not in [(c, t) for c, t in camiones_turnos]:
                camiones_turnos.append((camion, turno))
        
        for idx, (camion, turno) in enumerate(camiones_turnos):
            row_actual = 7 + idx
            
            # Nombre del camión en columna A
            ws.cell(row=row_actual, column=1).value = int(camion)
            
            # Turno en columna B (ahora muestra el turno específico, no todos)
            turno_display = str(turno) if turno and turno != 'N/A' else 'Sin turnos'
            ws.cell(row=row_actual, column=2).value = turno_display
            ws.cell(row=row_actual, column=2).alignment = center_alignment
            
            # Filtrar datos del camión Y TURNO específico
            df_camion_turno = df_transp[
                (df_transp['Camion'] == camion) & 
                (df_transp['Turno'] == turno)
            ].copy()
            df_camion_turno['fecha_key'] = pd.to_datetime(df_camion_turno['fecha_completa']).dt.date
            
            # Agrupar por fecha (en caso de que haya duplicados del mismo turno en el mismo día)
            df_camion_agrupado = df_camion_turno.groupby('fecha_key').agg({
                'Entregado_totalmente': 'sum',
                'Conexion_RCO': 'sum',
                'En_ruta': 'sum',
                'Planificado': 'sum'
            }).reset_index()
            
            # Crear diccionario fecha -> datos para acceso rápido
            datos_por_fecha = df_camion_agrupado.set_index('fecha_key').to_dict('index')
            
            # Llenar datos por fecha
            current_col = 3  # Empieza en columna C (después de Camion y Turno)
            for fecha in fechas:
                fecha_key = fecha.date()
                
                # Obtener datos del día (si existen)
                if fecha_key in datos_por_fecha:
                    datos_dia = datos_por_fecha[fecha_key]
                    entregado = datos_dia.get('Entregado_totalmente', 0)
                    conexion_rco = datos_dia.get('Conexion_RCO', 0)
                    en_ruta = datos_dia.get('En_ruta', 0)
                    planificado = datos_dia.get('Planificado', 0)
                    
                    # Criterio 1: Entrega realizada (Entregado_totalmente > 0)
                    criterio_1 = 1 if entregado > 0 else 0
                    
                    # Criterio 2: No entregado PERO con conexión RCO
                    criterio_2 = 1 if (entregado == 0 and conexion_rco > 0) else 0
                    
                    # Criterio 3: No entregado, sin RCO PERO con estadía en planta
                    criterio_3 = 1 if (entregado == 0 and conexion_rco == 0 and (en_ruta > 0 or planificado > 0)) else 0
                else:
                    # Sin datos para este día
                    criterio_1 = 0
                    criterio_2 = 0
                    criterio_3 = 0
                
                # Escribir valores
                ws.cell(row=row_actual, column=current_col).value = criterio_1
                ws.cell(row=row_actual, column=current_col + 1).value = criterio_2
                ws.cell(row=row_actual, column=current_col + 2).value = criterio_3
                
                current_col += 3
        
        # FILA DE TOTALES POR CRITERIO
        row_totales = 7 + len(camiones_turnos)
        
        # Etiqueta de totales
        cell_totales = ws.cell(row=row_totales, column=1)
        cell_totales.value = "Total"
        cell_totales.font = Font(bold=True)
        cell_totales.alignment = center_alignment
        
        # Dejar vacía la columna del turno
        ws.cell(row=row_totales, column=2).value = ""
        
        # Calcular totales por día y criterio
        current_col = 3  # Empieza en columna C
        for fecha in fechas:
            # Para cada día, sumar todos los camiones
            for offset_criterio in range(3):  # 3 criterios
                # Construir fórmula de suma para esta columna
                primera_fila_datos = 7
                ultima_fila_datos = 7 + len(camiones_turnos) - 1
                col_letter = openpyxl.utils.get_column_letter(current_col)
                
                formula = f"=SUM({col_letter}{primera_fila_datos}:{col_letter}{ultima_fila_datos})"
                cell_total = ws.cell(row=row_totales, column=current_col)
                cell_total.value = formula
                cell_total.font = Font(bold=True)
                cell_total.alignment = center_alignment
                
                current_col += 1
        
        # FILA DE TOTAL GENERAL (SUMA DE TODOS LOS CRITERIOS)
        row_total_general = row_totales + 1
        
        # Etiqueta
        cell_total_general = ws.cell(row=row_total_general, column=1)
        cell_total_general.value = "Total Criterios"
        cell_total_general.font = Font(bold=True)
        cell_total_general.alignment = center_alignment
        cell_total_general.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Dejar vacía la columna del turno
        ws.cell(row=row_total_general, column=2).value = ""
        ws.cell(row=row_total_general, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Calcular total general por día (suma de los 3 criterios)
        current_col = 3  # Empieza en columna C
        for fecha in fechas:
            # Sumar los 3 criterios del día
            col1_letter = openpyxl.utils.get_column_letter(current_col)
            col2_letter = openpyxl.utils.get_column_letter(current_col + 1)
            col3_letter = openpyxl.utils.get_column_letter(current_col + 2)
            
            formula = f"={col1_letter}{row_totales}+{col2_letter}{row_totales}+{col3_letter}{row_totales}"
            
            # Colocar el total en la primera columna del día y hacer merge de las 3 columnas
            ws.merge_cells(start_row=row_total_general, start_column=current_col,
                          end_row=row_total_general, end_column=current_col + 2)
            
            cell_general = ws.cell(row=row_total_general, column=current_col)
            cell_general.value = formula
            cell_general.font = Font(bold=True, size=11)
            cell_general.alignment = center_alignment
            cell_general.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            current_col += 3
        
        # FILA DE DISPONIBILIDAD (%)
        row_disponibilidad = row_total_general + 1
        
        # Etiqueta
        cell_disp_label = ws.cell(row=row_disponibilidad, column=1)
        cell_disp_label.value = "Disponibilidad (%)"
        cell_disp_label.font = Font(bold=True)
        cell_disp_label.alignment = center_alignment
        cell_disp_label.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        
        # Dejar vacía la columna del turno
        ws.cell(row=row_disponibilidad, column=2).value = ""
        ws.cell(row=row_disponibilidad, column=2).fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        
        # Calcular disponibilidad por día (Total Criterios / C2)
        current_col = 3  # Empieza en columna C
        for fecha in fechas:
            # Fórmula: Total Criterios del día / Banda (C2)
            col_total_letra = openpyxl.utils.get_column_letter(current_col)
            formula = f"=IF($C$2>0,{col_total_letra}{row_total_general}/$C$2,0)"
            
            # Colocar la disponibilidad en la primera columna del día y hacer merge de las 3 columnas
            ws.merge_cells(start_row=row_disponibilidad, start_column=current_col,
                          end_row=row_disponibilidad, end_column=current_col + 2)
            
            cell_disp = ws.cell(row=row_disponibilidad, column=current_col)
            cell_disp.value = formula
            cell_disp.font = Font(bold=True, size=11)
            cell_disp.alignment = center_alignment
            cell_disp.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            cell_disp.number_format = '0.00%'  # Formato de porcentaje
            
            current_col += 3
        
        # APLICAR BORDES Y AJUSTAR COLUMNAS
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12  # Columna Turno
        for col in range(3, current_col):  # Desde columna C en adelante
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[col_letter].width = 4
        
        # Aplicar bordes a la tabla de datos (incluye filas de totales y disponibilidad)
        max_data_row = 7 + len(camiones_turnos) + 2  # +3 para incluir TOTAL, TOTAL CRITERIOS y DISPONIBILIDAD
        for row in range(row_dia, max_data_row + 1):
            for col in range(1, current_col):
                ws.cell(row=row, column=col).border = thin_border
    
    # HOJAS ADICIONALES: LOGS (SOLO FLOTA LICITADA)
    
    # HOJA: LOG - Entregas Licitadas (CON TURNO)
    ws_entregas = workbook.create_sheet(title='LOG - Entregas Licitadas')
    entregas_data = df_licitados[df_licitados['Entregado_totalmente'] > 0][
        ['Fecha', 'Camion', 'Turno', 'Transportista', 'Entregado_totalmente']
    ].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(entregas_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_entregas.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Encabezado
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
    
    # HOJA: LOG - RCO Licitadas (CON TURNO)
    ws_rco = workbook.create_sheet(title='LOG - RCO Licitadas')
    rco_data = df_licitados[df_licitados['Conexion_RCO'] > 0][
        ['Fecha', 'Camion', 'Turno', 'Transportista', 'Conexion_RCO']
    ].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(rco_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_rco.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Encabezado
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
    
    # HOJA: LOG - Eventos Licitadas (CON TURNO)
    ws_eventos = workbook.create_sheet(title='LOG - Eventos Licitadas')
    eventos_data = df_licitados[
        ['Fecha', 'Camion', 'Turno', 'Transportista', 
         'Turnos_enviados', 'Conexion_RCO', 'Entregado_totalmente', 'En_ruta', 'Planificado']
    ].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(eventos_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_eventos.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Encabezado
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
    
    # HOJA ADICIONAL: OTROS TRANSPORTISTAS (NO LICITADOS)
    if not df_no_licitados.empty:
        ws_otros = workbook.create_sheet(title='OTROS Transportistas')
        
        # Agrupar datos de no licitados por transportista y fecha
        resumen_otros = df_no_licitados.groupby(['Transportista', 'Fecha']).agg({
            'Camion': 'count',  # Cantidad de camiones
            'Turnos_enviados': 'sum',
            'Conexion_RCO': 'sum',
            'Entregado_totalmente': 'sum'
        }).reset_index()
        resumen_otros.columns = ['Transportista', 'Fecha', 'Cantidad_Camiones', 
                                  'Turnos_enviados', 'Conexion_RCO', 'Entregado_totalmente']
        
        for r_idx, row in enumerate(dataframe_to_rows(resumen_otros, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_otros.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Encabezado
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
    
    # Guardar workbook
    workbook.save(output_buffer)


# ============================================
# RUTA PRINCIPAL: DETALLE (REFACTORIZADA)
# ============================================
@app.route('/detalle')
@auth.login_required
def detalle():
    """
    Vista de detalle con análisis granular por turno.
    """
    minera_nombre = request.args.get('minera')
    transportista_nombre = request.args.get('transportista')
    fecha = request.args.get('fecha')
    
    datos_completos_camiones = []
    banda_info = {
        'banda_total': 0,
        'bandas_por_transportista': {},
        'transportistas_unicos': [],
        'es_otro_transportista': False,
        'disponibilidad_operacional': {}
    }
    
    if USE_ATHENA and minera_nombre and transportista_nombre and fecha:
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        
        # OPTIMIZACIÓN: Cargar datos completos UNA SOLA VEZ usando caché
        datos_completos = obtener_datos_completos_athena(minera_nombre, fecha_obj, fecha_obj)
        
        if datos_completos is not None and not datos_completos.empty:            
            # Normalizar nombre del transportista para matching
            palabras_transportista = [
                palabra.upper().replace('.', '').replace(',', '')
                for palabra in transportista_nombre.split()
                if len(palabra) > 2
            ]
            
            def match_transportista(nombre_flota):
                if pd.isna(nombre_flota):
                    return False
                nombre_normalizado = nombre_flota.upper().replace('.', '').replace(',', '')
                return all(palabra in nombre_normalizado for palabra in palabras_transportista)
            
            print(f"🔍 DEBUG - Transportistas únicos en datos_completos:")
            print(datos_completos['Transportista'].unique())
            print(f"🔍 DEBUG - Buscando: {transportista_nombre}")
            print(f"🔍 DEBUG - Palabras clave: {palabras_transportista}")
            
            # Primero, identificar qué camiones pertenecen al transportista
            # (al menos un turno con actividad del transportista)
            camiones_del_transportista = datos_completos[
                datos_completos['Transportista'].apply(match_transportista)
            ]['Camion'].unique()
            
            print(f"🚛 Camiones identificados para {transportista_nombre}: {len(camiones_del_transportista)}")
            print(f"🚛 Camiones: {sorted(camiones_del_transportista)}")
            
            # Filtrar TODOS los turnos de esos camiones para esa fecha
            # Esto incluye turnos sin actividad
            df_transportista = datos_completos[
                datos_completos['Camion'].isin(camiones_del_transportista)
            ].copy()
            
            # Asegurar que el transportista esté correctamente asignado
            # (algunos turnos pueden tener SIN_INFORMACION si no tuvieron actividad)
            df_transportista['Transportista'] = df_transportista.apply(
                lambda row: transportista_nombre if match_transportista(row['Transportista']) 
                else (transportista_nombre if row['Transportista'] == 'SIN_INFORMACION' else row['Transportista']),
                axis=1
            )
            
            print(f"📊 Datos filtrados para {transportista_nombre}: {len(df_transportista)} registros (todos los turnos)")
            print(f"📊 Breakdown:")
            print(f"   - Con Entregado_totalmente > 0: {(df_transportista['Entregado_totalmente'] > 0).sum()}")
            print(f"   - Con Turnos_enviados > 0: {(df_transportista['Turnos_enviados'] > 0).sum()}")
            print(f"   - Con Conexion_RCO > 0: {(df_transportista['Conexion_RCO'] > 0).sum()}")
            
            if not df_transportista.empty:
                # Agregar columna 'Disponible' si no existe
                if 'Disponible' not in df_transportista.columns:
                    df_transportista['Disponible'] = (
                        (df_transportista['Entregado_totalmente'] > 0) | 
                        (df_transportista['Conexion_RCO'] > 0)
                    ).astype(int)
                
                # Renombrar columna para compatibilidad con template
                if '¿Es licitado?' in df_transportista.columns:
                    df_transportista['Es_licitado'] = df_transportista['¿Es licitado?']
                
                # Convertir a lista de diccionarios para template
                datos_completos_camiones = df_transportista.to_dict('records')
                
                # Calcular disponibilidad operacional
                banda_transportista = calcular_banda_transportista(transportista_nombre, minera_nombre)
                disponibilidad_operacional = calcular_disponibilidad_operacional(
                    df_transportista, 
                    banda_total=banda_transportista
                )
                
                banda_info = {
                    'banda_total': banda_transportista,
                    'bandas_por_transportista': {transportista_nombre: banda_transportista},
                    'transportistas_unicos': [transportista_nombre],
                    'es_otro_transportista': False,
                    'disponibilidad_operacional': disponibilidad_operacional
                }
    
    # Obtener información de flota licitada (TODOS los vehículos del mantenedor)
    flota_licitada = []
    if USE_ATHENA and minera_nombre and fecha:
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        
        # 1. Obtener TODOS los vehículos licitados del mantenedor
        df_mantenedor_flota = obtener_flota_licitada_athena(minera_nombre)
        
        if not df_mantenedor_flota.empty:
            # 2. Obtener datos de actividad (cacheados)
            datos_completos = obtener_datos_completos_athena(minera_nombre, fecha_obj, fecha_obj)
            
            # 3. Crear diccionario de actividad por camión (sumar todos los turnos del día)
            actividad_por_camion = {}
            if datos_completos is not None and not datos_completos.empty:
                # Agrupar por camión (sumar todos los turnos del día)
                actividad_agregada = datos_completos.groupby('Camion').agg({
                    'Turnos_enviados': 'sum',
                    'Conexion_RCO': 'sum',
                    'Entregado_totalmente': 'sum',
                    'Transportista': 'first'  # Tomar el primer transportista
                }).reset_index()
                
                for _, row in actividad_agregada.iterrows():
                    codigo = int(row['Camion'])
                    actividad_por_camion[codigo] = {
                        'turnos_enviados': int(row['Turnos_enviados']),
                        'conexiones_rco': int(row['Conexion_RCO']),
                        'entregas_realizadas': int(row['Entregado_totalmente']),
                        'transportista': row['Transportista']
                    }
            
            # 4. Crear registros combinando mantenedor con actividad
            registros_licitados = []
            for _, vehiculo in df_mantenedor_flota.iterrows():
                codigo_tanque = int(vehiculo['codigo_tanque'])
                
                # Obtener actividad (o defaults si no hay)
                actividad = actividad_por_camion.get(codigo_tanque, {
                    'turnos_enviados': 0,
                    'conexiones_rco': 0,
                    'entregas_realizadas': 0,
                    'transportista': vehiculo.get('nombre_transportista', 'SIN_INFORMACION')
                })
                
                registros_licitados.append({
                    'codigo_tanque': codigo_tanque,
                    'nombre_transportista': actividad['transportista'],
                    'turnos_enviados': actividad['turnos_enviados'],
                    'conexiones_rco': actividad['conexiones_rco'],
                    'entregas_realizadas': actividad['entregas_realizadas'],
                    'tiene_actividad': (actividad['entregas_realizadas'] > 0 or actividad['conexiones_rco'] > 0),
                    'uso': vehiculo.get('uso', ''),
                    'estado': 'Activo'  # Del mantenedor
                })
            
            flota_licitada = registros_licitados
            
            print(f"✅ Flota licitada completa: {len(flota_licitada)} vehículos totales")
            sin_actividad = len([v for v in flota_licitada if not v['tiene_actividad']])
            print(f"   - Con actividad: {len(flota_licitada) - sin_actividad}")
            print(f"   - Sin actividad: {sin_actividad}")
    
    mineras = obtener_mineras_athena()
    transportistas = obtener_transportistas_global() if USE_ATHENA else []

    return render_template('detalle.html', 
                         mineras=mineras,
                         minera=minera_nombre,
                         transportistas=transportistas,
                         transportista=transportista_nombre,
                         fecha=fecha,
                         datos_completos=datos_completos_camiones,
                         registros=datos_completos_camiones,
                         flota_licitada=flota_licitada,
                         banda_info=banda_info)


# ============================================
# OTRAS RUTAS Y ENDPOINTS
# ============================================

# Rutas principales
@app.route('/')
@auth.login_required
def index():
    """Vista principal - Dashboard (lista de mineras predefinidas)"""
    mineras = obtener_mineras_athena()
    data_source = 'Lista predefinida de mineras'
    return render_template('index.html', mineras=mineras, data_source=data_source)


@app.route('/exportar')
@auth.login_required
def exportar():
    """Vista para exportar datos a Excel"""
    mineras = obtener_mineras_athena()
    return render_template('exportar.html', mineras=mineras)

@app.route('/api/dashboard_data')
@auth.login_required
def dashboard_data():
    """API endpoint para obtener datos del dashboard (Athena-only)

    Query params expected:
    - minera: nombre de la minera (string)
    - mes: mes (int)
    - semana: semana dentro del mes (int)
    - año: año (int, opcional, por defecto año actual)
    Optional:
    - viajes_min: override mínimo de viajes
    - viajes_max: override máximo de viajes
    """
    minera_nombre = request.args.get('minera')
    mes = request.args.get('mes', type=int)
    semana = request.args.get('semana', type=int)
    año = request.args.get('año', type=int, default=datetime.now().year)
    
    # Obtener configuración específica para la minera
    config_viajes = obtener_configuracion_viajes()
    minera_config = config_viajes.get(minera_nombre, {'minimo': 11, 'maximo': 13})
    
    # Permitir override desde parámetros de la URL
    viajes_min = request.args.get('viajes_min', type=int, default=minera_config['minimo'])
    viajes_max = request.args.get('viajes_max', type=int, default=minera_config['maximo'])

    if not USE_ATHENA:
        return jsonify({'error': 'Athena no está disponible en este entorno'}), 503

    if not minera_nombre or not mes or not semana:
        return jsonify({'error': 'Parámetros incompletos'}), 400

    fecha_inicio, fecha_fin = calcular_rango_semana(año, mes, semana)

    grafico_data = obtener_datos_grafico_athena(minera_nombre, fecha_inicio, fecha_fin, viajes_min, viajes_max)
    matriz_data = obtener_datos_matriz_athena(minera_nombre, fecha_inicio, fecha_fin, viajes_min, viajes_max)

    # Las bandas ya están incluidas en matriz_data['bandas_transportistas']
    # No necesitamos calcularlas de nuevo
    bandas_por_transportista = matriz_data.get('bandas_transportistas', {})

    # Limpiar NaN antes de serializar a JSON
    response_data = {
        'grafico': grafico_data,
        'matriz': matriz_data,
        'bandas_transportista': bandas_por_transportista,
        'minera': {
            'nombre': minera_nombre,
            'viajes_minimos': viajes_min,
            'viajes_maximos': viajes_max
        },
        'source': 'Athena'
    }
    
    return jsonify(clean_nan_for_json(response_data))


@app.route('/api/semanas/<int:mes>')
@auth.login_required
def obtener_semanas(mes):
    """Obtener semanas disponibles para un mes"""
    año = request.args.get('año', type=int, default=datetime.now().year)
    semanas = calcular_semanas_mes(año, mes)
    return jsonify({'semanas': semanas})


def calcular_semanas_mes(año, mes):
    """Calcular número de semanas en un mes"""
    from calendar import monthrange
    dias_mes = monthrange(año, mes)[1]
    semanas = (dias_mes + 6) // 7
    return list(range(1, semanas + 1))


def calcular_rango_semana(año, mes, semana):
    """Calcular fecha de inicio y fin de una semana dentro de un mes"""
    primer_dia = datetime(año, mes, 1).date()
    dias_desde_inicio = (semana - 1) * 7
    fecha_inicio = primer_dia + timedelta(days=dias_desde_inicio)
    fecha_fin = fecha_inicio + timedelta(days=6)
    
    from calendar import monthrange
    ultimo_dia_mes = monthrange(año, mes)[1]
    fecha_fin_mes = datetime(año, mes, ultimo_dia_mes).date()
    
    if fecha_fin > fecha_fin_mes:
        fecha_fin = fecha_fin_mes
    
    return fecha_inicio, fecha_fin


# Removed SQLite-based helpers. Use Athena helpers: obtener_datos_grafico_athena and obtener_datos_matriz_athena

@app.route('/admin/mineras')
@auth.login_required
def admin_mineras():
    """Gestión de mineras y asociaciones"""
    mineras = obtener_mineras_athena()
    transportistas = obtener_transportistas_global() if USE_ATHENA else []
    return render_template('admin_mineras.html', mineras=mineras, transportistas=transportistas)


@app.route('/api/debug/transportistas')
@auth.login_required
def debug_transportistas():
    """Ruta temporal para identificar nombres exactos de transportistas"""
    if not USE_ATHENA:
        return jsonify({'error': 'Athena no está disponible'}), 503
    
    query = f"""
    SELECT DISTINCT carriername1 as transportista, COUNT(*) as total_registros
    FROM {DATABASE_ATHENA}.{TABLA_PEDIDOS}
    WHERE carriername1 IS NOT NULL 
    AND carriername1 != 'SIN_INFORMACION'
    GROUP BY carriername1
    ORDER BY total_registros DESC, transportista
    LIMIT 50
    """
    
    try:
        df = sql_athena(query)
        if df.empty:
            return jsonify({'transportistas': [], 'mensaje': 'No se encontraron transportistas'})
        
        # Convertir a lista de diccionarios
        transportistas_info = []
        for _, row in df.iterrows():
            transportistas_info.append({
                'nombre': row['transportista'],
                'registros': int(row['total_registros'])
            })
        
        return jsonify({
            'transportistas': transportistas_info,
            'total_encontrados': len(transportistas_info),
            'mensaje': 'Transportistas ordenados por cantidad de registros'
        })
        
    except Exception as e:
        return jsonify({'error': f'Error obteniendo transportistas: {str(e)}'}), 500


@app.route('/api/exportar_xlsx')
@auth.login_required
def exportar_xlsx():
    """Generar archivo XLSX con formato matricial para rango de fechas"""
    try:
        minera = request.args.get('minera')
        fecha_inicio_str = request.args.get('fecha_inicio')
        fecha_fin_str = request.args.get('fecha_fin')
        
        if not minera or not fecha_inicio_str or not fecha_fin_str:
            return jsonify({'error': 'Parámetros requeridos: minera, fecha_inicio, fecha_fin'}), 400
        
        # Convertir fechas
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
        
        if fecha_inicio > fecha_fin:
            return jsonify({'error': 'La fecha de inicio no puede ser mayor a la fecha de fin'}), 400
        
        # Obtener datos completos para el rango de fechas
        datos_completos = obtener_datos_completos_athena(minera, fecha_inicio, fecha_fin)
        
        if datos_completos is None or datos_completos.empty:
            return jsonify({'error': 'No se encontraron datos para los parámetros especificados'}), 404
        
        # Obtener vehículos licitados
        vehiculos_licitados = obtener_vehiculos_licitados_por_minera(minera)
        
        # Preparar DataFrame para exportación
        df_export = datos_completos.copy()
        
        # Agregar columna de vehículo licitado si no existe
        if '¿Es licitado?' in df_export.columns:
            df_export['Es_Licitado'] = df_export['¿Es licitado?']
        else:
            df_export['Es_Licitado'] = df_export['Camion'].isin(vehiculos_licitados).map({True: 'Si', False: 'No'})
        
        # Crear archivo Excel en memoria
        output = BytesIO()
        
        # Generar Excel con formato matricial
        exportar_xlsx_matricial(df_export, fecha_inicio, fecha_fin, output, minera)
        
        output.seek(0)
        
        # Preparar respuesta
        filename = f"DispoMinera_{minera}_{fecha_inicio_str}_{fecha_fin_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error generando XLSX: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error generando archivo: {str(e)}'}), 500


@app.route('/api/cache/clear', methods=['POST'])
@auth.login_required
def clear_cache():
    """Endpoint para limpiar el cache manualmente"""
    try:
        cache.clear()
        return jsonify({
            'success': True,
            'mensaje': 'Cache limpiado exitosamente'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/cache/stats', methods=['GET'])
@auth.login_required
def cache_stats():
    """Obtener estadísticas del cache (solo disponible con algunos backends)"""
    stats = {
        'cache_type': app.config['CACHE_TYPE'],
        'default_timeout': app.config['CACHE_DEFAULT_TIMEOUT'],
        'mensaje': 'Para estadísticas detalladas, use Redis como backend'
    }
    
    # Intentar obtener info adicional si es posible
    try:
        if hasattr(cache.cache, '_cache'):
            stats['cached_keys_count'] = len(cache.cache._cache)
            stats['cached_keys'] = list(cache.cache._cache.keys())[:10]  # Primeras 10
    except:
        pass
    
    return jsonify(stats)


@app.route('/api/mineras', methods=['GET', 'POST'])
@auth.login_required
def api_mineras():
    """Endpoints de mineras — Athena-only read API. POST no soportado."""
    if request.method == 'GET':
        mineras = obtener_mineras_athena()
        config_viajes = obtener_configuracion_viajes()
        
        # Devolver lista con configuración específica de cada minera
        return jsonify([{
            'nombre': m,
            'viajes_minimos': config_viajes.get(m, {'minimo': 11})['minimo'],
            'viajes_maximos': config_viajes.get(m, {'maximo': 13})['maximo']
        } for m in mineras])

    elif request.method == 'POST':
        return jsonify({'error': 'Creación de mineras no soportada en modo Athena-only'}), 405

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('FLASK_ENV', 'development') == 'development' 
    app.run(host='0.0.0.0', port=port, debug=debug)
